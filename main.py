"""
TelecomLens — main.py  (FastAPI backend, organisation-agnostic)
Run:  uvicorn main:app --reload --port 8000
Docs: http://localhost:8000/docs
"""
import os, re, sys, csv, io, subprocess, tempfile, hashlib, platform, json, logging, shutil
from pathlib import Path
from typing import Optional
from datetime import datetime

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
    datefmt="%H:%M:%S",
)
logger = logging.getLogger("telecomlens")

from fastapi import FastAPI, UploadFile, File, HTTPException, Depends, Query, Body
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse, HTMLResponse
from sqlalchemy import (
    create_engine, Column, String, Float, Boolean, Integer,
    DateTime, Text, JSON, func, text,
)
from sqlalchemy.orm import DeclarativeBase, Session, sessionmaker
from pydantic_settings import BaseSettings

from parser import parse_bill, classify_line
from discover import detect_org_tokens, classify_name, discover_patterns
from msisdn import normalise_msisdn, display_msisdn
# NOTE: report is imported lazily inside the endpoint so a missing python-docx
# does not crash the whole app at startup.



# ── Settings ──────────────────────────────────────────────────────────────────
class Settings(BaseSettings):
    database_url: str = "sqlite:///./telecomlens.db"
    poppler_path: str = "poppler"
    bills_folder: str = "bills"
    class Config:
        env_file = ".env"

settings = Settings()
STATIC_DIR = Path(__file__).parent / "static"

# ── Database ──────────────────────────────────────────────────────────────────
engine = create_engine(
    settings.database_url,
    connect_args={"check_same_thread": False} if "sqlite" in settings.database_url else {},
)
SessionLocal = sessionmaker(bind=engine)


def statement_to_iso(raw: str) -> str:
    """Convert a Safaricom 'dd/mm/yyyy' statement date into a sortable
    'YYYY-MM-DD' string. Returns '' if unparseable. Idempotent: a value that
    is already ISO is returned unchanged, so re-running the backfill is safe."""
    s = (raw or "").strip()
    if not s:
        return ""
    m = re.match(r"^(\d{4})-(\d{2})-(\d{2})$", s)          # already ISO
    if m:
        return s
    m = re.match(r"^(\d{1,2})/(\d{1,2})/(\d{4})$", s)      # dd/mm/yyyy
    if m:
        d, mo, y = m.groups()
        return f"{y}-{int(mo):02d}-{int(d):02d}"
    return ""


def _significant_digits(search: str):
    """Return the national significant digits of a phone-number-like search
    term, or None if it is not numeric. Canonical numbers are stored as
    254+national, so matching on these digits via ILIKE works for any input
    format (07…, 254…, +254…, spaced/hyphenated) and for partial fragments."""
    raw = (search or "").strip()
    if not raw or not re.fullmatch(r"[\d\s\-\+\(\)\.]+", raw):
        return None
    digits = re.sub(r"\D", "", raw)
    if len(digits) < 3:
        return None
    if digits.startswith("254") and len(digits) > 3:
        return digits[3:]
    if digits.startswith("0") and len(digits) > 1:
        return digits[1:]
    return digits


class Base(DeclarativeBase): pass

class Organisation(Base):
    __tablename__ = "organisations"
    id = Column(String, primary_key=True)
    name = Column(String)
    account_number = Column(String)
    created_at = Column(DateTime, default=datetime.utcnow)

class MappingRule(Base):
    __tablename__ = "mapping_rules"
    id = Column(Integer, primary_key=True, autoincrement=True)
    org_id = Column(String)
    pattern = Column(String)
    division = Column(String)
    priority = Column(Integer, default=0)

class GLAccount(Base):
    __tablename__ = "gl_accounts"
    id = Column(Integer, primary_key=True, autoincrement=True)
    org_id = Column(String)
    division = Column(String)
    gl_code = Column(String)

class BillUpload(Base):
    __tablename__ = "bill_uploads"
    id = Column(Integer, primary_key=True, autoincrement=True)
    org_id = Column(String)
    filename = Column(String)
    sha256 = Column(String, unique=True)
    statement_date = Column(String)
    statement_iso = Column(String, default="", index=True)   # YYYY-MM-DD, sortable
    account_total = Column(Float, default=0)
    outstanding_total = Column(Float, default=0)
    total_net = Column(Float, default=0)
    total_vat = Column(Float, default=0)
    total_excise = Column(Float, default=0)
    subscriber_count = Column(Integer, default=0)
    uploaded_at = Column(DateTime, default=datetime.utcnow)

class InvoiceLine(Base):
    __tablename__ = "invoice_lines"
    id = Column(Integer, primary_key=True, autoincrement=True)
    bill_id = Column(Integer)
    org_id = Column(String)
    invoice_number = Column(String)
    subscriber_number = Column(String)
    raw_name = Column(String)
    tariff_plan = Column(String)
    division = Column(String)
    geography = Column(String)
    pre_tax = Column(Float, default=0)
    excise = Column(Float, default=0)
    vat = Column(Float, default=0)
    amount_due_kes = Column(Float, default=0)
    outstanding = Column(Float, default=0)
    cdr_count = Column(Integer, default=0)
    is_anomaly = Column(Boolean, default=False)
    anomaly_reason = Column(String, default="")

class CDRRecord(Base):
    __tablename__ = "cdr_records"
    id = Column(Integer, primary_key=True, autoincrement=True)
    bill_id = Column(Integer)
    subscriber_number = Column(String)
    date = Column(String)
    time = Column(String)
    destination = Column(String)
    duration = Column(String)
    rate = Column(Float, default=0)
    charge = Column(Float, default=0)
    service_type = Column(String)


class SubscriberProfile(Base):
    """Persistent metadata for a subscriber line — survives across bill months."""
    __tablename__ = "subscriber_profiles"
    id = Column(Integer, primary_key=True, autoincrement=True)
    org_id = Column(String, index=True)
    subscriber_number = Column(String, index=True)
    display_name = Column(String, default="")          # human-readable alias
    division_override = Column(String, default="")     # manual cost-centre remap
    tags = Column(String, default="")                  # comma-separated tags
    device_type = Column(String, default="")           # Handset | SIM | Modem | Fixed
    tariff_override = Column(String, default="")       # admin-set expected plan
    notes = Column(String, default="")
    is_active = Column(Boolean, default=True)
    first_seen_date = Column(String, default="")
    last_seen_date = Column(String, default="")
    updated_at = Column(DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

class BudgetEntry(Base):
    """Monthly budget target per division for an org."""
    __tablename__ = "budget_entries"
    id = Column(Integer, primary_key=True, autoincrement=True)
    org_id = Column(String, index=True)
    division = Column(String)
    period = Column(String)   # "YYYY-MM"
    budget_kes = Column(Float, default=0)
    headcount = Column(Integer, default=0)

class SpendAlert(Base):
    """Per-subscriber or per-division spend ceiling."""
    __tablename__ = "spend_alerts"
    id = Column(Integer, primary_key=True, autoincrement=True)
    org_id = Column(String)
    scope = Column(String)          # "subscriber" | "division" | "total"
    scope_value = Column(String)    # subscriber number, division name, or ""
    threshold_kes = Column(Float, default=0)
    is_active = Column(Boolean, default=True)
    created_at = Column(DateTime, default=datetime.utcnow)


class Annotation(Base):
    """User comment attached to any invoice line or anomaly."""
    __tablename__ = "annotations"
    id         = Column(Integer, primary_key=True, autoincrement=True)
    org_id     = Column(String, index=True)
    bill_id    = Column(Integer, index=True)
    ref_type   = Column(String)          # "line" | "anomaly" | "division" | "bill"
    ref_id     = Column(String)          # subscriber_number, division name, or ""
    text       = Column(Text)
    author     = Column(String, default="")
    created_at = Column(DateTime, default=datetime.utcnow)

class AuditLog(Base):
    """Immutable record of every user action."""
    __tablename__ = "audit_logs"
    id         = Column(Integer, primary_key=True, autoincrement=True)
    org_id     = Column(String, index=True)
    action     = Column(String)          # e.g. "bill.upload", "subscriber.update"
    actor      = Column(String, default="user")
    detail     = Column(Text, default="")
    created_at = Column(DateTime, default=datetime.utcnow)

class WebhookConfig(Base):
    """Outbound webhook URL called after each bill import."""
    __tablename__ = "webhook_configs"
    id         = Column(Integer, primary_key=True, autoincrement=True)
    org_id     = Column(String)
    url        = Column(String)
    secret     = Column(String, default="")   # optional HMAC secret
    events     = Column(String, default="bill.imported")  # comma-separated
    is_active  = Column(Boolean, default=True)
    last_fired = Column(DateTime, nullable=True)


class Division(Base):
    """
    Org-specific division registry.  Stores the canonical list of division names
    the user has defined.  New names created during bulk-retag are auto-added here.
    """
    __tablename__ = "divisions"
    id         = Column(Integer, primary_key=True, autoincrement=True)
    org_id     = Column(String, index=True)
    name       = Column(String)                     # display name
    colour     = Column(String, default="")         # optional hex for charts
    description= Column(String, default="")
    created_at = Column(DateTime, default=datetime.utcnow)

class ChangeLog(Base):
    """
    Detailed, reversible change record.
    Every division retag or subscriber update writes a row here.
    The rollback endpoint reads prev_value to restore.
    """
    __tablename__ = "change_log"
    id            = Column(Integer, primary_key=True, autoincrement=True)
    org_id        = Column(String, index=True)
    entity_type   = Column(String)          # "invoice_line" | "subscriber_profile"
    entity_id     = Column(String)          # subscriber_number (stable across bills)
    field         = Column(String)          # "division" | "display_name" | "tags" etc.
    prev_value    = Column(Text, default="")
    new_value     = Column(Text, default="")
    actor         = Column(String, default="user")
    note          = Column(String, default="")
    created_at    = Column(DateTime, default=datetime.utcnow)
    rolled_back   = Column(Boolean, default=False)
    rolled_back_at= Column(DateTime, nullable=True)

Base.metadata.create_all(engine)


# ── Schema migration & one-time backfill ───────────────────────────────────────
# create_all() only creates *missing tables*; it does not add new columns to
# tables that already exist in a deployed DB. So new columns need an explicit,
# idempotent migration. (For non-SQLite engines, use a real migration tool.)

_BACKFILL_MARKER = "migrate.msisdn_iso.v1"


def _sqlite_column_exists(conn, table: str, col: str) -> bool:
    rows = conn.execute(text(f"PRAGMA table_info({table})")).fetchall()
    return any(r[1] == col for r in rows)


def _migrate_schema() -> None:
    """Add columns that create_all cannot add to pre-existing tables."""
    if "sqlite" not in settings.database_url:
        return
    with engine.begin() as conn:
        if not _sqlite_column_exists(conn, "bill_uploads", "statement_iso"):
            conn.execute(text(
                "ALTER TABLE bill_uploads ADD COLUMN statement_iso VARCHAR DEFAULT ''"
            ))
            logger.info("migrate: added bill_uploads.statement_iso")


def _merge_profiles(keep: "SubscriberProfile", dup: "SubscriberProfile") -> None:
    """Fold a duplicate profile (same canonical number) into the survivor."""
    di, ki = statement_to_iso(dup.first_seen_date), statement_to_iso(keep.first_seen_date)
    if di and (not ki or di < ki):
        keep.first_seen_date = dup.first_seen_date
    if statement_to_iso(dup.last_seen_date) > statement_to_iso(keep.last_seen_date):
        keep.last_seen_date = dup.last_seen_date
    for f in ("display_name", "division_override", "device_type", "tariff_override", "notes"):
        if not getattr(keep, f) and getattr(dup, f):
            setattr(keep, f, getattr(dup, f))
    tags = {t.strip() for t in (keep.tags or "").split(",") if t.strip()}
    tags |= {t.strip() for t in (dup.tags or "").split(",") if t.strip()}
    keep.tags = ",".join(sorted(tags))
    keep.is_active = bool(keep.is_active or dup.is_active)


def _backfill_once() -> None:
    """One-time: populate statement_iso and canonicalise every stored MSISDN.
    Guarded by an AuditLog marker so it runs once, then is a no-op each boot.
    The work is itself idempotent, so a crash mid-way is safely retried."""
    db = SessionLocal()
    try:
        if db.query(AuditLog).filter_by(action=_BACKFILL_MARKER).first():
            return
        logger.info("backfill: canonicalising MSISDNs + statement dates (one-time)…")

        for b in db.query(BillUpload).all():
            iso = statement_to_iso(b.statement_date)
            if iso and b.statement_iso != iso:
                b.statement_iso = iso

        for line in db.query(InvoiceLine).all():
            c = normalise_msisdn(line.subscriber_number)
            if c != (line.subscriber_number or ""):
                line.subscriber_number = c

        for cdr in db.query(CDRRecord).all():
            c = normalise_msisdn(cdr.subscriber_number)
            if c != (cdr.subscriber_number or ""):
                cdr.subscriber_number = c

        # Profiles: normalise, then merge any collisions the normalisation creates
        survivors: dict[tuple, SubscriberProfile] = {}
        for p in db.query(SubscriberProfile).all():
            p.subscriber_number = normalise_msisdn(p.subscriber_number)
            key = (p.org_id, p.subscriber_number)
            if key in survivors:
                _merge_profiles(survivors[key], p)
                db.delete(p)
            else:
                survivors[key] = p

        db.add(AuditLog(org_id="_system", action=_BACKFILL_MARKER,
                        detail="canonical MSISDN + statement_iso backfill complete"))
        db.commit()
        logger.info("backfill: complete")
    except Exception:
        db.rollback()
        logger.exception("backfill failed — will retry on next startup")
    finally:
        db.close()


_migrate_schema()
_backfill_once()


# ── App ───────────────────────────────────────────────────────────────────────
app = FastAPI(title="TelecomLens", version="4.2.0")
# CORS: wildcard is fine for a local single-machine tool.
# Set CORS_ORIGINS=http://myserver:8000 in .env to restrict in production.
_cors_origins = os.getenv("CORS_ORIGINS", "*").split(",")
app.add_middleware(CORSMiddleware, allow_origins=_cors_origins, allow_methods=["*"], allow_headers=["*"])


def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()


def find_pdftotext() -> str:
    if platform.system() == "Windows":
        local = Path(settings.poppler_path) / "Library" / "bin" / "pdftotext.exe"
        if local.exists():
            return str(local)
        for p in Path(settings.poppler_path).rglob("pdftotext.exe"):
            return str(p)
    path = shutil.which("pdftotext")
    if path:
        return path
    raise RuntimeError("pdftotext not found. Install poppler-utils (Linux) or run install.bat (Windows).")


def pdf_to_text(pdf_bytes: bytes) -> str:
    pdftotext = find_pdftotext()
    with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as f:
        f.write(pdf_bytes)
        tmp = f.name
    try:
        r = subprocess.run([pdftotext, "-layout", tmp, "-"],
                           capture_output=True, timeout=120)
        if r.returncode != 0:
            raise HTTPException(400, f"pdftotext error: {r.stderr.decode()[:500]}")
        return r.stdout.decode("utf-8", errors="replace")
    finally:
        Path(tmp).unlink(missing_ok=True)


def get_org_rules(org_id: str, db: Session) -> list[tuple[str, str]]:
    rows = db.query(MappingRule).filter_by(org_id=org_id).order_by(MappingRule.priority.desc()).all()
    return [(r.pattern, r.division) for r in rows]


def ensure_division(org_id: str, name: str, db: Session) -> None:
    """Create a division registry entry if it does not already exist."""
    if not name or not name.strip():
        return
    exists = db.query(Division).filter_by(org_id=org_id, name=name.strip()).first()
    if not exists:
        db.add(Division(org_id=org_id, name=name.strip()))
        db.flush()


def _log_change(
    org_id: str,
    entity_type: str,
    entity_id: str,
    field: str,
    prev_value: str,
    new_value: str,
    actor: str = "user",
    note: str = "",
    db: Session = None,
) -> "ChangeLog":
    """Write one ChangeLog row and return it."""
    entry = ChangeLog(
        org_id=org_id,
        entity_type=entity_type,
        entity_id=entity_id,
        field=field,
        prev_value=str(prev_value) if prev_value is not None else "",
        new_value=str(new_value) if new_value is not None else "",
        actor=actor,
        note=note,
    )
    db.add(entry)
    return entry


def detect_anomalies(inv: dict) -> tuple[bool, str]:
    reasons = []
    if inv.get("amount_due_kes", 0) == 0 and inv.get("cdr_count", 0) > 0:
        reasons.append("zero charge with activity")
    if inv.get("amount_due_kes", 0) > 50000:
        reasons.append("high spend >50K")
    if inv.get("cdr_count", 0) > 1000:
        reasons.append("high CDR count")
    if inv.get("division", "").startswith("Other") or inv.get("division") == "Unclassified":
        reasons.append("unclassified line")
    return bool(reasons), "; ".join(reasons)


def store_bill(bill_data: dict, filename: str, sha: str, db: Session) -> BillUpload:
    org_id = re.sub(r"[^a-z0-9]", "_", bill_data["org_account"].lower()) or "default"
    org = db.query(Organisation).filter_by(id=org_id).first()
    if not org:
        org = Organisation(id=org_id, name=bill_data["org_name"],
                           account_number=bill_data["org_account"])
        db.add(org)
    rules = get_org_rules(org_id, db)
    bill = BillUpload(
        org_id=org_id, filename=filename, sha256=sha,
        statement_date=bill_data["statement_date"],
        statement_iso=statement_to_iso(bill_data["statement_date"]),
        account_total=bill_data["account_total"],
        outstanding_total=bill_data.get("outstanding", 0),
        total_net=bill_data.get("total_net", 0),
        total_vat=bill_data.get("total_vat", 0),
        total_excise=bill_data.get("total_excise", 0),
        subscriber_count=len(bill_data["invoices"]),
    )
    db.add(bill)
    db.flush()
    # Canonicalise subscriber numbers up front so invoice lines, CDRs and
    # profiles below all key off the same value across bills.
    for inv in bill_data["invoices"]:
        inv["subscriber_number"] = normalise_msisdn(inv.get("subscriber_number"))
    for inv in bill_data["invoices"]:
        existing = db.query(InvoiceLine).filter_by(invoice_number=inv["invoice_number"]).first()
        if existing:
            continue
        is_anom, reason = detect_anomalies(inv)
        line = InvoiceLine(
            bill_id=bill.id, org_id=org_id,
            invoice_number=inv["invoice_number"],
            subscriber_number=inv["subscriber_number"],
            raw_name=inv["raw_name"],
            tariff_plan=inv["tariff_plan"],
            division=classify_line(
                raw_name=inv["raw_name"],
                tariff_plan=inv.get("tariff_plan", ""),
                cdrs=inv.get("cdr_records"),
                amount_due=inv.get("amount_due_kes", 0),
                user_rules=rules or None,
            ),
            geography=inv["geography"],
            pre_tax=inv["pre_tax"], excise=inv["excise"],
            vat=inv["vat"], amount_due_kes=inv["amount_due_kes"],
            outstanding=inv["outstanding"], cdr_count=inv["cdr_count"],
            is_anomaly=is_anom, anomaly_reason=reason,
        )
        db.add(line)
        for cdr in inv.get("cdr_records", [])[:5000]:
            db.add(CDRRecord(
                bill_id=bill.id, subscriber_number=inv["subscriber_number"],
                date=cdr["date"], time=cdr["time"], destination=cdr["destination"],
                duration=cdr["duration"], rate=cdr["rate"], charge=cdr["charge"],
                service_type=cdr["service_type"],
            ))
    # ── Upsert subscriber profiles (lifecycle tracking) ────────────────
    for inv in bill_data["invoices"]:
        if not inv.get("subscriber_number"):
            continue
        sub_num = inv["subscriber_number"]
        profile = db.query(SubscriberProfile).filter_by(
            org_id=org_id, subscriber_number=sub_num).first()
        if not profile:
            profile = SubscriberProfile(
                org_id=org_id,
                subscriber_number=sub_num,
                display_name=inv.get("raw_name", ""),
                first_seen_date=bill_data.get("statement_date", ""),
                last_seen_date=bill_data.get("statement_date", ""),
                is_active=True,
            )
            db.add(profile)
        else:
            # Update first/last seen by chronological (ISO) order, so imports
            # in any order still yield a correct first-introduced date.
            sd = bill_data.get("statement_date", "")
            sd_iso = statement_to_iso(sd)
            if sd_iso and sd_iso > statement_to_iso(profile.last_seen_date or ""):
                profile.last_seen_date = sd
            if sd_iso and (not statement_to_iso(profile.first_seen_date)
                           or sd_iso < statement_to_iso(profile.first_seen_date)):
                profile.first_seen_date = sd
            profile.is_active = True

    db.commit()

    # Fire webhooks asynchronously (best-effort)
    _fire_webhooks_bg(org_id, bill, db)

    # Audit log
    db.add(AuditLog(org_id=org_id, action="bill.import",
                    detail=f"Imported {filename}, {len(bill_data['invoices'])} lines"))
    db.commit()
    return bill


def _fire_webhooks_bg(org_id: str, bill: "BillUpload", db: Session):
    """Best-effort synchronous webhook call (runs in same thread, non-blocking on error)."""
    hooks = db.query(WebhookConfig).filter_by(org_id=org_id, is_active=True).all()
    if not hooks:
        return
    import threading, json as _json, hmac as _hmac, hashlib as _hl
    payload = _json.dumps({
        "event": "bill.imported",
        "org_id": org_id,
        "bill_id": bill.id,
        "statement_date": bill.statement_date,
        "account_total": bill.account_total,
        "subscriber_count": bill.subscriber_count,
    }).encode()
    def _send(hook):
        try:
            import httpx
            headers = {"Content-Type": "application/json", "X-TelecomLens-Event": "bill.imported"}
            if hook.secret:
                sig = _hmac.new(hook.secret.encode(), payload, _hl.sha256).hexdigest()
                headers["X-TelecomLens-Signature"] = f"sha256={sig}"
            httpx.post(hook.url, content=payload, headers=headers, timeout=8)
            hook.last_fired = datetime.utcnow()
            db.commit()
            logger.info("Webhook fired: %s → %s", org_id, hook.url)
        except Exception as exc:
            logger.warning("Webhook failed (%s): %s", hook.url, exc)
    for hook in hooks:
        threading.Thread(target=_send, args=(hook,), daemon=True).start()


# ── Endpoints ─────────────────────────────────────────────────────────────────

@app.get("/api/health")
def health():
    try:
        pt = find_pdftotext(); ok = True
    except RuntimeError as e:
        pt = str(e); ok = False
    return {"status": "ok", "platform": platform.system(), "pdftotext_found": ok,
            "pdftotext_path": pt}


@app.get("/api/orgs")
def list_orgs(db: Session = Depends(get_db)):
    orgs = db.query(Organisation).all()
    return [{"id": o.id, "name": o.name, "account_number": o.account_number} for o in orgs]


@app.get("/api/bills")
def list_bills(org_id: Optional[str] = None, db: Session = Depends(get_db)):
    q = db.query(BillUpload)
    if org_id:
        q = q.filter_by(org_id=org_id)
    bills = q.order_by(BillUpload.statement_iso.desc()).all()
    return [{"id": b.id, "org_id": b.org_id, "filename": b.filename,
             "statement_date": b.statement_date, "account_total": b.account_total,
             "outstanding_total": getattr(b, "outstanding_total", 0) or 0,
             "subscriber_count": b.subscriber_count,
             "uploaded_at": b.uploaded_at.isoformat() if b.uploaded_at else ""} for b in bills]


@app.post("/api/bills/upload")
async def upload_bill(file: UploadFile = File(...), db: Session = Depends(get_db)):
    data = await file.read()
    sha = hashlib.sha256(data).hexdigest()
    existing = db.query(BillUpload).filter_by(sha256=sha).first()
    if existing:
        return {"status": "duplicate", "bill_id": existing.id}
    logger.info("Extracting text from PDF (%d bytes)…", len(data))
    text = pdf_to_text(data)
    logger.info("Parsing %d chars of bill text…", len(text))
    bill_data = parse_bill(text)
    logger.info("Storing %d invoices…", len(bill_data["invoices"]))
    bill = store_bill(bill_data, file.filename or "upload.pdf", sha, db)
    logger.info("Import complete: bill_id=%d, org=%s, subscribers=%d",
                bill.id, bill.org_id, bill.subscriber_count)
    return {
        "status": "ok",
        "bill_id": bill.id,
        "org_id": bill.org_id,
        "subscriber_count": bill.subscriber_count,
        "account_total": bill_data["account_total"],
        "statement_date": bill_data["statement_date"],
        "org_name": bill_data["org_name"],
    }


@app.post("/api/bills/import-folder")
def import_folder(folder: str = Body(..., embed=True), db: Session = Depends(get_db)):
    p = Path(folder).resolve()
    if not p.exists():
        raise HTTPException(404, f"Folder not found: {folder}")
    results = []
    for pdf in sorted(p.glob("*.pdf")):
        data = pdf.read_bytes()
        sha = hashlib.sha256(data).hexdigest()
        existing = db.query(BillUpload).filter_by(sha256=sha).first()
        if existing:
            results.append({"file": pdf.name, "status": "duplicate", "bill_id": existing.id})
            continue
        try:
            text = pdf_to_text(data)
            bill_data = parse_bill(text)
            bill = store_bill(bill_data, pdf.name, sha, db)
            results.append({"file": pdf.name, "status": "ok", "bill_id": bill.id})
        except Exception as e:
            results.append({"file": pdf.name, "status": "error", "error": str(e)[:200]})
    return results


@app.get("/api/bills/{bill_id}/summary")
def bill_summary(bill_id: int, db: Session = Depends(get_db)):
    bill = db.query(BillUpload).filter_by(id=bill_id).first()
    if not bill:
        raise HTTPException(404, "Bill not found")
    lines = db.query(InvoiceLine).filter_by(bill_id=bill_id).all()
    # Use bill-level totals (parsed directly from TAX ANALYSIS section = exact)
    # Fall back to summing lines if bill-level fields not populated (older imports)
    total       = bill.account_total if bill.account_total else sum(l.amount_due_kes for l in lines)
    pre_tax     = bill.total_net     if bill.total_net     else sum(l.pre_tax for l in lines)
    excise      = bill.total_excise  if bill.total_excise  else sum(l.excise for l in lines)
    vat         = bill.total_vat     if bill.total_vat     else sum(l.vat for l in lines)
    outstanding = bill.outstanding_total if bill.outstanding_total else sum(l.outstanding for l in lines)
    anomaly_count = sum(1 for l in lines if l.is_anomaly)
    divisions = {}
    for l in lines:
        divisions[l.division] = divisions.get(l.division, 0) + l.amount_due_kes
    top_division = max(divisions, key=divisions.get) if divisions else ""
    return {
        "bill_id": bill_id, "org_id": bill.org_id, "filename": bill.filename,
        "statement_date": bill.statement_date, "account_total": round(total, 2),
        "pre_tax_total": round(pre_tax, 2), "excise_total": round(excise, 2),
        "vat_total": round(vat, 2), "outstanding_total": round(outstanding, 2),
        "subscriber_count": len(lines), "anomaly_count": anomaly_count,
        "top_division": top_division,
        "divisions": {k: round(v, 2) for k, v in sorted(divisions.items(), key=lambda x: -x[1])},
    }


@app.get("/api/bills/{bill_id}/divisions")
def bill_divisions(bill_id: int, db: Session = Depends(get_db)):
    rows = (db.query(InvoiceLine.division,
                     func.sum(InvoiceLine.amount_due_kes).label("total"),
                     func.count(InvoiceLine.id).label("count"))
            .filter_by(bill_id=bill_id)
            .group_by(InvoiceLine.division)
            .order_by(func.sum(InvoiceLine.amount_due_kes).desc()).all())
    return [{"division": r.division, "total": round(r.total or 0, 2), "count": r.count} for r in rows]


@app.get("/api/bills/{bill_id}/subscribers")
def bill_subscribers(bill_id: int, search: str = "", division: str = "",
                     page: int = 1, limit: int = 50, db: Session = Depends(get_db)):
    q = db.query(InvoiceLine).filter_by(bill_id=bill_id)
    if search:
        safe_search = search.replace("\\", "\\\\").replace("%", "\\%").replace("_", "\\_")
        q = q.filter(InvoiceLine.raw_name.ilike(f"%{safe_search}%") |
                     InvoiceLine.subscriber_number.ilike(f"%{safe_search}%"))
    if division:
        q = q.filter_by(division=division)
    total = q.count()
    rows = q.order_by(InvoiceLine.amount_due_kes.desc()).offset((page - 1) * limit).limit(limit).all()
    return {
        "total": total, "page": page, "limit": limit,
        "subscribers": [{
            "subscriber_number": r.subscriber_number, "raw_name": r.raw_name,
            "tariff_plan": r.tariff_plan, "division": r.division, "geography": r.geography,
            "amount_due_kes": round(r.amount_due_kes, 2), "pre_tax": round(r.pre_tax, 2),
            "excise": round(r.excise, 2), "vat": round(r.vat, 2),
            "cdr_count": r.cdr_count, "is_anomaly": r.is_anomaly,
            "anomaly_reason": r.anomaly_reason,
        } for r in rows],
    }


@app.get("/api/bills/{bill_id}/subscriber/{sub_number}/cdr")
def subscriber_cdr(bill_id: int, sub_number: str, limit: int = 500,
                   db: Session = Depends(get_db)):
    rows = (db.query(CDRRecord)
            .filter_by(bill_id=bill_id, subscriber_number=normalise_msisdn(sub_number))
            .order_by(CDRRecord.date, CDRRecord.time)
            .limit(min(limit, 5000)).all())
    return [{
        "date": r.date, "time": r.time, "destination": r.destination,
        "duration": r.duration, "rate": r.rate,
        "charge": round(r.charge, 2), "service_type": r.service_type,
    } for r in rows]


@app.get("/api/bills/{bill_id}/anomalies")
def bill_anomalies(bill_id: int, db: Session = Depends(get_db)):
    rows = db.query(InvoiceLine).filter_by(bill_id=bill_id, is_anomaly=True).all()
    return [{
        "subscriber_number": r.subscriber_number, "raw_name": r.raw_name,
        "division": r.division, "amount_due_kes": round(r.amount_due_kes, 2),
        "anomaly_reason": r.anomaly_reason, "cdr_count": r.cdr_count,
    } for r in rows]


@app.get("/api/bills/{bill_id}/drilldown")
def bill_drilldown(bill_id: int,
                   by: str = Query(..., description="division|subscriber|tariff|anomaly|geography|unclassified"),
                   value: str = Query("", description="Value to filter on"),
                   db: Session = Depends(get_db)):
    """Return line items for a specific field/value combination — powers the drill-down panel."""
    q = db.query(InvoiceLine).filter_by(bill_id=bill_id)
    if by == "division":
        q = q.filter_by(division=value)
    elif by == "subscriber":
        q = q.filter_by(subscriber_number=value)
    elif by == "tariff":
        safe_val = value.replace("\\", "\\\\").replace("%", "\\%").replace("_", "\\_")
        q = q.filter(InvoiceLine.tariff_plan.ilike(f"%{safe_val}%"))
    elif by == "anomaly":
        q = q.filter_by(is_anomaly=True)
    elif by == "geography":
        q = q.filter_by(geography=value)
    elif by == "unclassified":
        q = q.filter(
            (InvoiceLine.division == "Unclassified")
            | (InvoiceLine.division.ilike("Other%"))
        )

    rows = q.order_by(InvoiceLine.amount_due_kes.desc()).limit(200).all()
    total_amount = sum(r.amount_due_kes for r in rows)
    total_count = len(rows)

    # Monthly history for sparkline (reuses across all bills in same org)
    org_id = db.query(BillUpload.org_id).filter_by(id=bill_id).scalar()
    history = []
    if org_id and by == "division" and value:
        bills = db.query(BillUpload).filter_by(org_id=org_id).order_by(BillUpload.statement_iso).all()
        for b in bills:
            amt = (db.query(func.sum(InvoiceLine.amount_due_kes))
                   .filter_by(bill_id=b.id, division=value).scalar() or 0)
            history.append({"date": b.statement_date, "amount": round(float(amt), 2)})

    return {
        "by": by, "value": value, "total_amount": round(total_amount, 2),
        "count": total_count, "history": history,
        "lines": [{
            "subscriber_number": r.subscriber_number,
            "raw_name": r.raw_name,
            "tariff_plan": r.tariff_plan,
            "division": r.division,
            "geography": r.geography,
            "pre_tax": round(r.pre_tax, 2),
            "excise": round(r.excise, 2),
            "vat": round(r.vat, 2),
            "amount_due_kes": round(r.amount_due_kes, 2),
            "cdr_count": r.cdr_count,
            "is_anomaly": r.is_anomaly,
            "anomaly_reason": r.anomaly_reason,
        } for r in rows],
    }


@app.get("/api/bills/{bill_id}/chargeback.csv")
def chargeback_csv(bill_id: int, db: Session = Depends(get_db)):
    bill = db.query(BillUpload).filter_by(id=bill_id).first()
    if not bill:
        raise HTTPException(404)
    gl_map = {g.division: g.gl_code for g in db.query(GLAccount).filter_by(org_id=bill.org_id).all()}
    rows = db.query(InvoiceLine).filter_by(bill_id=bill_id).order_by(InvoiceLine.division, InvoiceLine.raw_name).all()
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(["Subscriber Number", "Name", "Division", "GL Account",
                "Tariff Plan", "Pre-Tax (KES)", "Excise (KES)", "VAT (KES)",
                "Amount Due (KES)", "Outstanding (KES)", "Anomaly"])
    for r in rows:
        w.writerow([r.subscriber_number, r.raw_name, r.division,
                    gl_map.get(r.division, ""), r.tariff_plan,
                    r.pre_tax, r.excise, r.vat, r.amount_due_kes,
                    r.outstanding, "Yes" if r.is_anomaly else ""])
    buf.seek(0)
    return StreamingResponse(iter([buf.getvalue()]),
                             media_type="text/csv",
                             headers={"Content-Disposition": f'attachment; filename="chargeback_{bill_id}.csv"'})


@app.get("/api/trends")
def trends(org_id: Optional[str] = None, db: Session = Depends(get_db)):
    q = db.query(BillUpload)
    if org_id:
        q = q.filter_by(org_id=org_id)
    bills = q.order_by(BillUpload.statement_iso).all()
    result = []
    for b in bills:
        lines = db.query(InvoiceLine).filter_by(bill_id=b.id).all()
        anomalies = sum(1 for l in lines if l.is_anomaly)
        divisions = {}
        for l in lines:
            divisions[l.division] = round((divisions.get(l.division, 0) + l.amount_due_kes), 2)
        result.append({
            "bill_id": b.id, "statement_date": b.statement_date,
            "account_total": round(b.account_total, 2),
            "subscriber_count": b.subscriber_count,
            "anomaly_count": anomalies,
            "divisions": divisions,
        })
    return result


@app.get("/api/trends/top-spenders")
def top_spenders(org_id: Optional[str] = None, limit: int = 15, db: Session = Depends(get_db)):
    q_bills = db.query(BillUpload)
    if org_id:
        q_bills = q_bills.filter_by(org_id=org_id)
    bills = q_bills.order_by(BillUpload.statement_iso).all()
    bill_map = {b.id: b.statement_date for b in bills}
    org_filter = InvoiceLine.org_id == org_id if org_id else True
    totals = (db.query(InvoiceLine.subscriber_number, InvoiceLine.raw_name, InvoiceLine.division,
                       func.sum(InvoiceLine.amount_due_kes).label("total"),
                       func.count(InvoiceLine.bill_id.distinct()).label("month_count"))
              .filter(org_filter)
              .group_by(InvoiceLine.subscriber_number, InvoiceLine.raw_name, InvoiceLine.division)
              .order_by(func.sum(InvoiceLine.amount_due_kes).desc()).limit(limit).all())
    result = []
    for t in totals:
        hist = (db.query(InvoiceLine.bill_id, InvoiceLine.amount_due_kes, InvoiceLine.tariff_plan)
                .filter_by(subscriber_number=t.subscriber_number)
                .order_by(InvoiceLine.bill_id).all())
        result.append({
            "subscriber_number": t.subscriber_number, "raw_name": t.raw_name,
            "division": t.division, "total_all_months": round(t.total or 0, 2),
            "month_count": t.month_count,
            "history": [{"statement_date": bill_map.get(r.bill_id, ""),
                         "amount_kes": r.amount_due_kes, "tariff_plan": r.tariff_plan}
                        for r in hist],
        })
    return result


@app.get("/api/orgs/{org_id}/rules")
def get_rules(org_id: str, db: Session = Depends(get_db)):
    rules = db.query(MappingRule).filter_by(org_id=org_id).all()
    return [{"id": r.id, "pattern": r.pattern, "division": r.division, "priority": r.priority} for r in rules]


@app.post("/api/orgs/{org_id}/rules")
def add_rule(org_id: str, body: dict = Body(...), db: Session = Depends(get_db)):
    rule = MappingRule(org_id=org_id, pattern=body["pattern"],
                       division=body["division"], priority=body.get("priority", 0))
    db.add(rule); db.commit()
    return {"id": rule.id}


@app.delete("/api/orgs/{org_id}/rules/{rule_id}")
def delete_rule(org_id: str, rule_id: int, db: Session = Depends(get_db)):
    db.query(MappingRule).filter_by(id=rule_id, org_id=org_id).delete()
    db.commit()
    return {"ok": True}


@app.get("/api/bills/{bill_id}/report.docx")
def download_report(bill_id: int, db: Session = Depends(get_db)):
    """Generate and return a .docx executive report for the given bill."""
    bill = db.query(BillUpload).filter_by(id=bill_id).first()
    if not bill:
        raise HTTPException(404, "Bill not found")

    org = db.query(Organisation).filter_by(id=bill.org_id).first()
    lines = db.query(InvoiceLine).filter_by(bill_id=bill_id).all()

    total       = sum(l.amount_due_kes for l in lines)
    pre_tax     = sum(l.pre_tax        for l in lines)
    excise      = sum(l.excise         for l in lines)
    vat         = sum(l.vat            for l in lines)
    outstanding = sum(l.outstanding    for l in lines)
    anomaly_count = sum(1 for l in lines if l.is_anomaly)

    divisions_raw = {}
    division_counts = {}
    for l in lines:
        divisions_raw[l.division] = divisions_raw.get(l.division, 0) + l.amount_due_kes
        division_counts[l.division] = division_counts.get(l.division, 0) + 1
    divisions = [{"division": k, "total": round(v, 2), "count": division_counts[k]}
                 for k, v in sorted(divisions_raw.items(), key=lambda x: -x[1])]
    top_division = divisions[0]["division"] if divisions else ""

    anomalies = [{"subscriber_number": l.subscriber_number, "raw_name": l.raw_name,
                  "division": l.division, "amount_due_kes": round(l.amount_due_kes, 2),
                  "anomaly_reason": l.anomaly_reason, "cdr_count": l.cdr_count}
                 for l in lines if l.is_anomaly]

    top_subscribers = [{"subscriber_number": l.subscriber_number, "raw_name": l.raw_name,
                        "division": l.division, "tariff_plan": l.tariff_plan,
                        "amount_due_kes": round(l.amount_due_kes, 2)}
                       for l in sorted(lines, key=lambda x: -x.amount_due_kes)[:15]]

    # Multi-month trend from same org
    trend_bills = (db.query(BillUpload).filter_by(org_id=bill.org_id)
                   .order_by(BillUpload.statement_iso).all())
    trends = []
    for b in trend_bills:
        b_lines = db.query(InvoiceLine).filter_by(bill_id=b.id).all()
        trends.append({
            "statement_date":   b.statement_date,
            "account_total":    round(sum(l.amount_due_kes for l in b_lines), 2),
            "subscriber_count": len(b_lines),
            "anomaly_count":    sum(1 for l in b_lines if l.is_anomaly),
        })

    report_data = {
        "org_name":       org.name if org else bill.org_id,
        "account_number": org.account_number if org else "",
        "statement_date": bill.statement_date,
        "top_division":   top_division,
        "summary": {
            "account_total":    round(total, 2),
            "pre_tax_total":    round(pre_tax, 2),
            "excise_total":     round(excise, 2),
            "vat_total":        round(vat, 2),
            "outstanding_total":round(outstanding, 2),
            "subscriber_count": len(lines),
            "anomaly_count":    anomaly_count,
        },
        "divisions":        divisions,
        "anomalies":        anomalies,
        "top_subscribers":  top_subscribers,
        "trends":           trends,
        "waste":            _compute_waste(bill.org_id, db),
    }

    try:
        from report import generate_report
    except ImportError:
        raise HTTPException(500,
            "python-docx is not installed. Run: pip install python-docx --break-system-packages")

    try:
        docx_bytes = generate_report(report_data)
    except Exception as e:
        raise HTTPException(500, f"Report generation failed: {e}")

    safe_date = (bill.statement_date or "").replace("/", "-").replace(" ", "_")
    filename  = f"TelecomLens_{bill.org_id}_{safe_date}.docx"
    return StreamingResponse(
        io.BytesIO(docx_bytes),
        media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )





@app.post("/api/bills/{bill_id}/reclassify")
def reclassify_bill(bill_id: int, db: Session = Depends(get_db)):
    """
    Re-run multi-signal classification on all lines in a bill.
    Useful after upgrading the parser or adding user rules.
    Does NOT overwrite lines that have a subscriber profile with division_override set.
    """
    bill = db.query(BillUpload).filter_by(id=bill_id).first()
    if not bill:
        raise HTTPException(404, "Bill not found")

    rules = get_org_rules(bill.org_id, db)
    lines = db.query(InvoiceLine).filter_by(bill_id=bill_id).all()

    # Build CDR map for service-mix classification
    cdr_map: dict[str, list[dict]] = {}
    cdrs = db.query(CDRRecord).filter_by(bill_id=bill_id).all()
    for c in cdrs:
        cdr_map.setdefault(c.subscriber_number, []).append({
            "service_type": c.service_type, "charge": c.charge,
        })

    # Build override map from subscriber profiles
    overrides: dict[str, str] = {}
    profiles = db.query(SubscriberProfile).filter_by(org_id=bill.org_id).all()
    for p in profiles:
        if p.division_override:
            overrides[p.subscriber_number] = p.division_override

    changed = 0
    for line in lines:
        # Skip lines with a manual division override
        if line.subscriber_number in overrides:
            if line.division != overrides[line.subscriber_number]:
                line.division = overrides[line.subscriber_number]
                changed += 1
            continue

        new_div = classify_line(
            raw_name=line.raw_name or "",
            tariff_plan=line.tariff_plan or "",
            cdrs=cdr_map.get(line.subscriber_number),
            amount_due=line.amount_due_kes or 0,
            user_rules=rules or None,
        )
        if new_div != line.division:
            line.division = new_div
            changed += 1

        # Also update anomaly flag
        is_anom, reason = detect_anomalies({
            "amount_due_kes": line.amount_due_kes,
            "cdr_count": line.cdr_count,
            "division": new_div,
        })
        line.is_anomaly = is_anom
        line.anomaly_reason = reason

    db.commit()
    logger.info("Reclassified bill %d: %d/%d lines changed", bill_id, changed, len(lines))
    db.add(AuditLog(org_id=bill.org_id, action="bill.reclassify",
                    detail=f"bill={bill_id}, {changed}/{len(lines)} changed"))
    db.commit()
    return {"bill_id": bill_id, "total_lines": len(lines), "changed": changed}

# ═══════════════════════════════════════════════════════════════════════════════
# REPORTING & EXPORT ENDPOINTS
# ═══════════════════════════════════════════════════════════════════════════════

@app.post("/api/bills/{bill_id}/report-custom.docx")
def download_custom_report(bill_id: int, body: dict = Body(...), db: Session = Depends(get_db)):
    """
    Custom report builder. body fields (all optional, default True):
      include_summary, include_tax, include_divisions, include_subscribers,
      include_anomalies, include_trends, include_annotations,
      divisions_filter: [list of division names to include, empty = all],
      top_n_subscribers: int (default 15)
    """
    bill = db.query(BillUpload).filter_by(id=bill_id).first()
    if not bill:
        raise HTTPException(404, "Bill not found")

    org = db.query(Organisation).filter_by(id=bill.org_id).first()
    lines_q = db.query(InvoiceLine).filter_by(bill_id=bill_id)

    div_filter = body.get("divisions_filter", [])
    if div_filter:
        lines_q = lines_q.filter(InvoiceLine.division.in_(div_filter))
    lines = lines_q.all()

    annotations = db.query(Annotation).filter_by(bill_id=bill_id).order_by(Annotation.created_at).all()
    ann_map: dict[str, list] = {}
    for a in annotations:
        ann_map.setdefault(a.ref_id, []).append({"text": a.text, "author": a.author,
                                                  "created_at": str(a.created_at)[:16]})

    total       = sum(l.amount_due_kes for l in lines)
    pre_tax     = sum(l.pre_tax for l in lines)
    excise      = sum(l.excise for l in lines)
    vat         = sum(l.vat for l in lines)
    outstanding = sum(l.outstanding for l in lines)

    divs_raw: dict[str, float] = {}
    div_counts: dict[str, int] = {}
    for l in lines:
        divs_raw[l.division] = divs_raw.get(l.division, 0) + l.amount_due_kes
        div_counts[l.division] = div_counts.get(l.division, 0) + 1
    divisions = [{"division": k, "total": round(v, 2), "count": div_counts[k]}
                 for k, v in sorted(divs_raw.items(), key=lambda x: -x[1])]

    top_n = int(body.get("top_n_subscribers", 15))
    top_subscribers = [
        {"subscriber_number": l.subscriber_number, "raw_name": l.raw_name,
         "division": l.division, "tariff_plan": l.tariff_plan,
         "amount_due_kes": round(l.amount_due_kes, 2),
         "annotations": ann_map.get(l.subscriber_number, [])}
        for l in sorted(lines, key=lambda x: -x.amount_due_kes)[:top_n]
    ]
    anomalies = [
        {"subscriber_number": l.subscriber_number, "raw_name": l.raw_name,
         "division": l.division, "amount_due_kes": round(l.amount_due_kes, 2),
         "anomaly_reason": l.anomaly_reason, "cdr_count": l.cdr_count,
         "annotations": ann_map.get(l.subscriber_number, [])}
        for l in lines if l.is_anomaly
    ]

    trend_bills = (db.query(BillUpload).filter_by(org_id=bill.org_id)
                   .order_by(BillUpload.statement_iso).all())
    trends = []
    for b in trend_bills:
        bl = db.query(InvoiceLine).filter_by(bill_id=b.id).all()
        trends.append({"statement_date": b.statement_date,
                       "account_total": round(sum(x.amount_due_kes for x in bl), 2),
                       "subscriber_count": len(bl),
                       "anomaly_count": sum(1 for x in bl if x.is_anomaly)})

    report_data = {
        "org_name": org.name if org else bill.org_id,
        "account_number": org.account_number if org else "",
        "statement_date": bill.statement_date,
        "top_division": divisions[0]["division"] if divisions else "",
        "summary": {
            "account_total": round(total, 2), "pre_tax_total": round(pre_tax, 2),
            "excise_total": round(excise, 2), "vat_total": round(vat, 2),
            "outstanding_total": round(outstanding, 2),
            "subscriber_count": len(lines),
            "anomaly_count": sum(1 for l in lines if l.is_anomaly),
        },
        "divisions": divisions if body.get("include_divisions", True) else [],
        "anomalies": anomalies if body.get("include_anomalies", True) else [],
        "top_subscribers": top_subscribers if body.get("include_subscribers", True) else [],
        "trends": trends if body.get("include_trends", True) else [],
        "waste": _compute_waste(bill.org_id, db) if body.get("include_waste", True) else {},
        "options": {
            "include_summary": body.get("include_summary", True),
            "include_tax": body.get("include_tax", True),
            "include_annotations": body.get("include_annotations", True),
            "division_filter_label": ", ".join(div_filter) if div_filter else "All divisions",
        },
    }

    try:
        from report import generate_report
    except ImportError:
        raise HTTPException(500, "python-docx not installed. Run: pip install python-docx")
    try:
        docx_bytes = generate_report(report_data)
    except Exception as e:
        raise HTTPException(500, f"Report generation failed: {e}")

    safe = (bill.statement_date or "").replace("/", "-").replace(" ", "_")
    suffix = "_custom" if div_filter else ""
    filename = f"TelecomLens_{bill.org_id}_{safe}{suffix}.docx"

    db.add(AuditLog(org_id=bill.org_id, action="report.export",
                    detail=f"Custom report bill={bill_id}, divs={div_filter or 'all'}"))
    db.commit()

    return StreamingResponse(io.BytesIO(docx_bytes),
        media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'})


@app.get("/api/bills/{bill_id}/chargeback-excel.xlsx")
def chargeback_excel(bill_id: int, db: Session = Depends(get_db)):
    """
    Multi-sheet Excel workbook: one sheet per division + a Summary sheet.
    Each sheet has the subscriber lines for that division.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(500, "openpyxl not installed. Run: pip install openpyxl")

    bill = db.query(BillUpload).filter_by(id=bill_id).first()
    if not bill:
        raise HTTPException(404)

    gl_map = {g.division: g.gl_code
              for g in db.query(GLAccount).filter_by(org_id=bill.org_id).all()}
    lines = (db.query(InvoiceLine).filter_by(bill_id=bill_id)
             .order_by(InvoiceLine.division, InvoiceLine.amount_due_kes.desc()).all())
    annotations = db.query(Annotation).filter_by(bill_id=bill_id).all()
    ann_map = {a.ref_id: a.text for a in annotations}

    # Group lines by division
    from collections import defaultdict
    by_div: dict[str, list] = defaultdict(list)
    for l in lines:
        by_div[l.division].append(l)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    BLUE_FILL = PatternFill("solid", fgColor="2563EB")
    LIGHT_FILL = PatternFill("solid", fgColor="DBEAFE")
    HDR_FONT  = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    BODY_FONT = Font(name="Calibri", size=10)
    ANOM_FILL = PatternFill("solid", fgColor="FEE2E2")
    thin = Side(style="thin", color="CCCCCC")
    BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

    COLS = ["Subscriber Number", "Name", "Division", "GL Account", "Tariff Plan",
            "Pre-Tax (KES)", "Excise (KES)", "VAT (KES)", "Amount Due (KES)",
            "Outstanding (KES)", "CDR Count", "Anomaly", "Notes"]
    WIDTHS = [18, 28, 16, 12, 24, 14, 12, 12, 16, 16, 10, 8, 30]

    def _write_sheet(ws, rows_data, title_label):
        ws.freeze_panes = "A2"
        # Header
        for col_i, (hdr, w) in enumerate(zip(COLS, WIDTHS), start=1):
            cell = ws.cell(1, col_i, hdr)
            cell.font = HDR_FONT
            cell.fill = BLUE_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = BORDER
            ws.column_dimensions[get_column_letter(col_i)].width = w
        ws.row_dimensions[1].height = 20
        # Data
        for ri, l in enumerate(rows_data, start=2):
            is_anom = l.is_anomaly
            vals = [l.subscriber_number, l.raw_name, l.division,
                    gl_map.get(l.division, ""), l.tariff_plan or "",
                    round(l.pre_tax, 2), round(l.excise, 2), round(l.vat, 2),
                    round(l.amount_due_kes, 2), round(l.outstanding, 2),
                    l.cdr_count, "Yes" if is_anom else "",
                    ann_map.get(l.subscriber_number, "")]
            fill = ANOM_FILL if is_anom else (LIGHT_FILL if ri % 2 == 0 else None)
            for ci, v in enumerate(vals, start=1):
                cell = ws.cell(ri, ci, v)
                cell.font = BODY_FONT
                cell.border = BORDER
                if fill:
                    cell.fill = fill
        # Totals row
        total_row = len(rows_data) + 2
        ws.cell(total_row, 1, "TOTAL").font = Font(bold=True, size=10)
        for ci, col_name in enumerate(COLS, start=1):
            if col_name in ("Pre-Tax (KES)", "Excise (KES)", "VAT (KES)",
                            "Amount Due (KES)", "Outstanding (KES)"):
                col_ltr = get_column_letter(ci)
                ws.cell(total_row, ci,
                        f"=SUM({col_ltr}2:{col_ltr}{total_row-1})"
                        ).font = Font(bold=True, size=10)

    # Per-division sheets
    for div in sorted(by_div.keys()):
        safe_name = re.sub(r'[\\/:*?<>|]', '_', div)[:31]
        ws = wb.create_sheet(title=safe_name)
        _write_sheet(ws, by_div[div], div)

    # Summary sheet (first)
    ws_sum = wb.create_sheet(title="Summary", index=0)
    ws_sum.freeze_panes = "A2"
    sum_cols = ["Division", "GL Account", "Subscribers",
                "Amount Due (KES)", "% of Total", "Anomalies"]
    sum_widths = [20, 14, 12, 18, 12, 10]
    for ci, (h, w) in enumerate(zip(sum_cols, sum_widths), start=1):
        c = ws_sum.cell(1, ci, h)
        c.font = HDR_FONT; c.fill = BLUE_FILL
        c.alignment = Alignment(horizontal="center"); c.border = BORDER
        ws_sum.column_dimensions[get_column_letter(ci)].width = w
    grand_total = sum(l.amount_due_kes for l in lines) or 1
    for ri, div in enumerate(sorted(by_div.keys()), start=2):
        div_lines = by_div[div]
        div_total = sum(l.amount_due_kes for l in div_lines)
        div_anom  = sum(1 for l in div_lines if l.is_anomaly)
        vals = [div, gl_map.get(div, ""), len(div_lines),
                round(div_total, 2), f"{div_total/grand_total*100:.1f}%", div_anom]
        for ci, v in enumerate(vals, start=1):
            c = ws_sum.cell(ri, ci, v)
            c.font = BODY_FONT; c.border = BORDER
            if ri % 2 == 0:
                c.fill = LIGHT_FILL
    # Grand total
    tr = len(by_div) + 2
    ws_sum.cell(tr, 1, "GRAND TOTAL").font = Font(bold=True, size=10)
    ws_sum.cell(tr, 3, len(lines)).font = Font(bold=True)
    ws_sum.cell(tr, 4, round(sum(l.amount_due_kes for l in lines), 2)).font = Font(bold=True)
    ws_sum.cell(tr, 6, sum(1 for l in lines if l.is_anomaly)).font = Font(bold=True)

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)

    safe_date = (bill.statement_date or "").replace("/", "-").replace(" ", "_")
    filename = f"TelecomLens_Chargeback_{bill.org_id}_{safe_date}.xlsx"

    db.add(AuditLog(org_id=bill.org_id, action="report.excel",
                    detail=f"Excel chargeback bill={bill_id}"))
    db.commit()

    return StreamingResponse(buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'})


# ─── Annotation endpoints ─────────────────────────────────────────────────────

@app.get("/api/bills/{bill_id}/annotations")
def get_annotations(bill_id: int, ref_id: str = "", db: Session = Depends(get_db)):
    q = db.query(Annotation).filter_by(bill_id=bill_id)
    if ref_id:
        q = q.filter_by(ref_id=ref_id)
    rows = q.order_by(Annotation.created_at).all()
    return [{"id": r.id, "ref_type": r.ref_type, "ref_id": r.ref_id,
             "text": r.text, "author": r.author,
             "created_at": str(r.created_at)[:16]} for r in rows]


@app.post("/api/bills/{bill_id}/annotations")
def add_annotation(bill_id: int, body: dict = Body(...), db: Session = Depends(get_db)):
    bill = db.query(BillUpload).filter_by(id=bill_id).first()
    if not bill:
        raise HTTPException(404)
    ann = Annotation(
        org_id=bill.org_id, bill_id=bill_id,
        ref_type=body.get("ref_type", "line"),
        ref_id=body.get("ref_id", ""),
        text=body.get("text", "").strip(),
        author=body.get("author", "user"),
    )
    db.add(ann); db.commit()
    db.add(AuditLog(org_id=bill.org_id, action="annotation.add",
                    detail=f"bill={bill_id} ref={body.get('ref_id','')}"))
    db.commit()
    return {"id": ann.id}


@app.delete("/api/bills/{bill_id}/annotations/{ann_id}")
def delete_annotation(bill_id: int, ann_id: int, db: Session = Depends(get_db)):
    ann = db.query(Annotation).filter_by(id=ann_id, bill_id=bill_id).first()
    if not ann:
        raise HTTPException(404)
    bill = db.query(BillUpload).filter_by(id=bill_id).first()
    db.delete(ann); db.commit()
    if bill:
        db.add(AuditLog(org_id=bill.org_id, action="annotation.delete",
                        detail=f"bill={bill_id} ann={ann_id}"))
        db.commit()
    return {"ok": True}


# ═══════════════════════════════════════════════════════════════════════════════
# DATA & INTEGRATIONS ENDPOINTS
# ═══════════════════════════════════════════════════════════════════════════════

@app.get("/api/orgs/{org_id}/audit-log")
def get_audit_log(
    org_id: str,
    limit: int = 100,
    action: str = "",
    db: Session = Depends(get_db),
):
    q = db.query(AuditLog).filter_by(org_id=org_id)
    if action:
        q = q.filter(AuditLog.action.ilike(f"%{action}%"))
    rows = q.order_by(AuditLog.created_at.desc()).limit(min(limit, 500)).all()
    return [{"id": r.id, "action": r.action, "actor": r.actor,
             "detail": r.detail, "created_at": str(r.created_at)[:19]} for r in rows]


@app.get("/api/orgs/{org_id}/webhooks")
def list_webhooks(org_id: str, db: Session = Depends(get_db)):
    rows = db.query(WebhookConfig).filter_by(org_id=org_id).all()
    return [{"id": r.id, "url": r.url, "events": r.events,
             "is_active": r.is_active,
             "last_fired": str(r.last_fired)[:16] if r.last_fired else None} for r in rows]


@app.post("/api/orgs/{org_id}/webhooks")
def add_webhook(org_id: str, body: dict = Body(...), db: Session = Depends(get_db)):
    wh = WebhookConfig(
        org_id=org_id,
        url=body["url"],
        secret=body.get("secret", ""),
        events=body.get("events", "bill.imported"),
        is_active=True,
    )
    db.add(wh); db.commit()
    db.add(AuditLog(org_id=org_id, action="webhook.add", detail=f"url={body['url']}"))
    db.commit()
    return {"id": wh.id}


@app.delete("/api/orgs/{org_id}/webhooks/{wh_id}")
def delete_webhook(org_id: str, wh_id: int, db: Session = Depends(get_db)):
    db.query(WebhookConfig).filter_by(id=wh_id, org_id=org_id).delete()
    db.add(AuditLog(org_id=org_id, action="webhook.delete", detail=f"id={wh_id}"))
    db.commit()
    return {"ok": True}


@app.post("/api/orgs/{org_id}/webhooks/{wh_id}/test")
def test_webhook(org_id: str, wh_id: int, db: Session = Depends(get_db)):
    """Send a test ping to a webhook URL."""
    hook = db.query(WebhookConfig).filter_by(id=wh_id, org_id=org_id).first()
    if not hook:
        raise HTTPException(404)
    try:
        import httpx, json as _json
        payload = _json.dumps({"event": "ping", "org_id": org_id, "test": True}).encode()
        r = httpx.post(hook.url, content=payload,
                       headers={"Content-Type": "application/json",
                                "X-TelecomLens-Event": "ping"},
                       timeout=8)
        return {"status": r.status_code, "ok": r.is_success}
    except Exception as e:
        raise HTTPException(400, f"Webhook test failed: {e}")


@app.get("/api/health/carriers")
def detect_carrier(db: Session = Depends(get_db)):
    """
    Return which carriers have been detected across all imported bills,
    based on account number patterns and org names.
    """
    orgs = db.query(Organisation).all()
    results = []
    for org in orgs:
        carrier = _detect_carrier_from_org(org)
        results.append({"org_id": org.id, "org_name": org.name, "carrier": carrier})
    return results


def _detect_carrier_from_org(org) -> str:
    """Infer carrier from account number prefix or org name patterns."""
    name = (org.name or "").upper()
    acc  = (org.account_number or "").upper()
    if re.search(r"SAFARICOM|SAF\b|\bSAF\d", acc) or "SAFARICOM" in name:
        return "Safaricom"
    if re.search(r"AIRTEL|AIR\b", acc) or "AIRTEL" in name:
        return "Airtel Kenya"
    if re.search(r"TELKOM|TKL\b", acc) or "TELKOM" in name:
        return "Telkom Kenya"
    if re.search(r"FAIBA|JTL\b", acc) or "FAIBA" in name:
        return "Faiba / JTL"
    return "Unknown"




@app.get("/api/health/status")
def server_status():
    """Quick ping to verify server is alive (used by UI progress tracking)."""
    return {"ok": True, "ts": datetime.utcnow().isoformat()}

# ═══════════════════════════════════════════════════════════════════════════════
# SUBSCRIBER MANAGEMENT ENDPOINTS
# ═══════════════════════════════════════════════════════════════════════════════

@app.get("/api/orgs/{org_id}/subscribers")
def list_subscriber_profiles(
    org_id: str,
    search: str = "",
    tag: str = "",
    division: str = "",
    page: int = 1,
    limit: int = 60,
    db: Session = Depends(get_db),
):
    """All known subscriber profiles for an org, with latest bill data joined."""
    q = db.query(SubscriberProfile).filter_by(org_id=org_id)
    if search:
        safe = search.replace("\\", "\\\\").replace("%", "\\%").replace("_", "\\_")
        sig = _significant_digits(search)
        if sig:
            # Phone search: match canonical numbers by national significant digits,
            # so 07…, 254…, +254…, spaced, and partial inputs all resolve.
            q = q.filter(SubscriberProfile.subscriber_number.ilike(f"%{sig}%"))
        else:
            # Also match against raw_name stored on any InvoiceLine for this subscriber
            matched_subs = (
                db.query(InvoiceLine.subscriber_number)
                .filter(
                    InvoiceLine.org_id == org_id,
                    InvoiceLine.raw_name.ilike(f"%{safe}%"),
                )
                .distinct()
                .subquery()
            )
            q = q.filter(
                SubscriberProfile.subscriber_number.ilike(f"%{safe}%")
                | SubscriberProfile.display_name.ilike(f"%{safe}%")
                | SubscriberProfile.subscriber_number.in_(matched_subs)
            )
    if tag:
        safe_tag = tag.replace("\\", "\\\\").replace("%", "\\%").replace("_", "\\_")
        q = q.filter(SubscriberProfile.tags.ilike(f"%{safe_tag}%"))
    if division:
        # Match on override OR actual division from latest line
        sub_with_div = (
            db.query(InvoiceLine.subscriber_number)
            .filter(InvoiceLine.org_id == org_id, InvoiceLine.division == division)
            .distinct()
            .subquery()
        )
        q = q.filter(
            (SubscriberProfile.division_override == division)
            | SubscriberProfile.subscriber_number.in_(sub_with_div)
        )
    total = q.count()
    profiles = q.order_by(SubscriberProfile.display_name).offset((page - 1) * limit).limit(limit).all()

    # Join latest bill spend
    result = []
    for p in profiles:
        latest = (
            db.query(InvoiceLine)
            .join(BillUpload, InvoiceLine.bill_id == BillUpload.id)
            .filter(InvoiceLine.org_id == org_id,
                    InvoiceLine.subscriber_number == p.subscriber_number)
            .order_by(BillUpload.statement_iso.desc())
            .first()
        )
        result.append({
            "subscriber_number": p.subscriber_number,
            "display_number": display_msisdn(p.subscriber_number),
            "display_name": p.display_name,
            "division_override": p.division_override,
            "division": p.division_override or (latest.division if latest else ""),
            "tags": p.tags,
            "device_type": p.device_type,
            "tariff_override": p.tariff_override,
            "notes": p.notes,
            "is_active": p.is_active,
            "first_seen_date": p.first_seen_date,
            "last_seen_date": p.last_seen_date,
            "latest_amount_kes": round(latest.amount_due_kes, 2) if latest else 0,
            "latest_tariff": latest.tariff_plan if latest else "",
            "is_anomaly": latest.is_anomaly if latest else False,
        })
    return {"total": total, "page": page, "limit": limit, "profiles": result}


@app.patch("/api/orgs/{org_id}/subscribers/{sub_number}")
def update_subscriber_profile(
    org_id: str,
    sub_number: str,
    body: dict = Body(...),
    db: Session = Depends(get_db),
):
    """Update display_name, division_override, tags, device_type, notes, tariff_override."""
    sub_number = normalise_msisdn(sub_number)
    profile = db.query(SubscriberProfile).filter_by(
        org_id=org_id, subscriber_number=sub_number
    ).first()
    if not profile:
        # Auto-create if missing (e.g. manual add before bill import)
        profile = SubscriberProfile(org_id=org_id, subscriber_number=sub_number)
        db.add(profile)
    ALLOWED = {"display_name", "division_override", "tags", "device_type",
               "notes", "tariff_override", "is_active"}
    actor = body.pop("actor", "user")
    note  = body.pop("note", "")
    for k, v in body.items():
        if k in ALLOWED:
            prev = getattr(profile, k, "")
            if str(prev) != str(v):
                _log_change(org_id, "subscriber_profile", sub_number,
                            k, prev, v, actor=actor, note=note, db=db)
            setattr(profile, k, v)
    # If division_override changed, ensure it exists in the division registry
    if "division_override" in body and body["division_override"]:
        ensure_division(org_id, body["division_override"], db)
    profile.updated_at = datetime.utcnow()
    db.add(AuditLog(org_id=org_id, action="subscriber.update",
                    detail=f"sub={sub_number} fields={list(body.keys())}"))
    db.commit()
    return {"ok": True}


# cost-change thresholds for derived spend events
_SPIKE_PCT = 0.5            # ±50%
_SPIKE_FLOOR_KES = 100.0    # ignore swings smaller than this in absolute terms


def _msisdn_variants(canonical: str) -> set[str]:
    """Formats a number may have been stored in before canonicalisation, so
    ChangeLog rows written pre-normalisation still match the canonical key."""
    v = {canonical}
    if len(canonical) == 12 and canonical.startswith("254"):
        v |= {"0" + canonical[3:], "+" + canonical, canonical[3:]}
    return v


@app.get("/api/orgs/{org_id}/subscribers/{sub_number}/history")
def subscriber_history(org_id: str, sub_number: str, as_of: str = "",
                       db: Session = Depends(get_db)):
    """Full cross-bill history for one subscriber number: per-period billing
    rows, derived lifecycle events (from presence-stable bill fields), and the
    manual change/retag audit trail (from ChangeLog).

    `as_of` (optional, 'YYYY-MM' or 'YYYY-MM-DD') pins the reference period:
    only bills up to and including it are analysed, so a stray partial or
    out-of-cycle bill imported for a later date does not flip status/lifecycle.
    Defaults to the most recent bill."""
    canonical = normalise_msisdn(sub_number)

    all_bills = (db.query(BillUpload).filter_by(org_id=org_id)
                 .order_by(BillUpload.statement_iso).all())
    available_periods = [{"statement_date": b.statement_date,
                          "statement_iso": b.statement_iso} for b in all_bills]

    as_of = (as_of or "").strip()
    if as_of:
        if re.match(r"^\d{4}-\d{2}$", as_of):          # whole month, inclusive
            bills = [b for b in all_bills if (b.statement_iso or "")[:7] <= as_of]
        else:                                          # specific date, inclusive
            bills = [b for b in all_bills if (b.statement_iso or "") <= as_of]
    else:
        bills = all_bills

    # A number can carry more than one invoice line in a single bill — aggregate.
    agg: dict[int, dict] = {}
    for l in (db.query(InvoiceLine)
              .filter_by(org_id=org_id, subscriber_number=canonical).all()):
        a = agg.get(l.bill_id)
        if a is None:
            a = agg[l.bill_id] = {
                "raw_name": l.raw_name or "", "division": l.division or "",
                "tariff_plan": l.tariff_plan or "", "amount_due_kes": 0.0,
                "pre_tax": 0.0, "vat": 0.0, "excise": 0.0, "cdr_count": 0,
                "is_anomaly": False, "anomaly_reason": "", "_max_amt": -1.0,
            }
        a["amount_due_kes"] += l.amount_due_kes or 0
        a["pre_tax"] += l.pre_tax or 0
        a["vat"] += l.vat or 0
        a["excise"] += l.excise or 0
        a["cdr_count"] += l.cdr_count or 0
        a["is_anomaly"] = a["is_anomaly"] or bool(l.is_anomaly)
        if l.anomaly_reason and l.anomaly_reason not in a["anomaly_reason"]:
            a["anomaly_reason"] = (a["anomaly_reason"] + "; " + l.anomaly_reason).strip("; ")
        # representative name/division/tariff = the dominant (largest) line
        if (l.amount_due_kes or 0) > a["_max_amt"]:
            a["_max_amt"] = l.amount_due_kes or 0
            a["raw_name"] = l.raw_name or a["raw_name"]
            a["division"] = l.division or a["division"]
            a["tariff_plan"] = l.tariff_plan or a["tariff_plan"]

    # Chronological per-period timeline with running deltas
    timeline = []
    prev_amt = None
    for b in bills:
        if b.id not in agg:
            continue
        a = agg[b.id]
        amt = round(a["amount_due_kes"], 2)
        delta = None if prev_amt is None else round(amt - prev_amt, 2)
        delta_pct = (round((amt - prev_amt) / abs(prev_amt) * 100, 1)
                     if prev_amt not in (None, 0) else None)
        timeline.append({
            "bill_id": b.id, "statement_date": b.statement_date,
            "statement_iso": b.statement_iso, "raw_name": a["raw_name"],
            "division": a["division"], "tariff_plan": a["tariff_plan"],
            "amount_due_kes": amt, "pre_tax": round(a["pre_tax"], 2),
            "vat": round(a["vat"], 2), "excise": round(a["excise"], 2),
            "cdr_count": a["cdr_count"], "is_anomaly": a["is_anomaly"],
            "anomaly_reason": a["anomaly_reason"],
            "delta_kes": delta, "delta_pct": delta_pct,
        })
        prev_amt = amt

    # Derived lifecycle events — only from fields that are NOT retroactively
    # rewritten (name, tariff, presence, amount). Division history deliberately
    # comes from the ChangeLog audit trail below, not from diffing lines.
    events = []
    present = [i for i, b in enumerate(bills) if b.id in agg]
    if present:
        fb = bills[present[0]]
        events.append({"type": "first_seen", "statement_date": fb.statement_date,
                        "statement_iso": fb.statement_iso,
                        "detail": f"First appeared in {fb.statement_date}"})
        for ai, bi in zip(present, present[1:]):
            pa, pb, cur = agg[bills[ai].id], agg[bills[bi].id], bills[bi]
            if bi - ai > 1:
                events.append({"type": "reactivated", "statement_date": cur.statement_date,
                               "statement_iso": cur.statement_iso,
                               "detail": f"Reappeared after absent from {bi - ai - 1} bill(s)"})
            if pb["raw_name"] and pa["raw_name"] != pb["raw_name"]:
                events.append({"type": "name_change", "statement_date": cur.statement_date,
                               "statement_iso": cur.statement_iso,
                               "from": pa["raw_name"], "to": pb["raw_name"],
                               "detail": f"Name on bill changed: {pa['raw_name']} → {pb['raw_name']}"})
            if pa["tariff_plan"] and pb["tariff_plan"] and pa["tariff_plan"] != pb["tariff_plan"]:
                events.append({"type": "plan_change", "statement_date": cur.statement_date,
                               "statement_iso": cur.statement_iso,
                               "from": pa["tariff_plan"], "to": pb["tariff_plan"],
                               "detail": f"Tariff changed: {pa['tariff_plan']} → {pb['tariff_plan']}"})
            pamt, camt = round(pa["amount_due_kes"], 2), round(pb["amount_due_kes"], 2)
            if pamt > 0 and abs(camt - pamt) >= _SPIKE_FLOOR_KES:
                pct = (camt - pamt) / pamt
                if pct >= _SPIKE_PCT:
                    events.append({"type": "cost_spike", "statement_date": cur.statement_date,
                                   "statement_iso": cur.statement_iso, "from": pamt, "to": camt,
                                   "detail": f"Spend rose {round(pct*100)}% ({pamt:,.0f} → {camt:,.0f} KES)"})
                elif pct <= -_SPIKE_PCT:
                    events.append({"type": "cost_drop", "statement_date": cur.statement_date,
                                   "statement_iso": cur.statement_iso, "from": pamt, "to": camt,
                                   "detail": f"Spend fell {round(abs(pct)*100)}% ({pamt:,.0f} → {camt:,.0f} KES)"})
        last_i = present[-1]
        if last_i < len(bills) - 1:
            gb = bills[last_i + 1]
            events.append({"type": "gone", "statement_date": gb.statement_date,
                           "statement_iso": gb.statement_iso,
                           "detail": f"Absent from {gb.statement_date} onward "
                                     f"(last seen {bills[last_i].statement_date})"})
    events.sort(key=lambda e: e.get("statement_iso") or "", reverse=True)

    # Status from presence in the most recent bill
    status = "unknown"
    if present:
        last_bill = bills[-1]
        if last_bill.id in agg:
            r = agg[last_bill.id]
            status = "dormant" if (round(r["amount_due_kes"], 2) == 0
                                   and r["cdr_count"] == 0) else "active"
        else:
            status = "gone"

    profile = db.query(SubscriberProfile).filter_by(
        org_id=org_id, subscriber_number=canonical).first()

    lifetime = round(sum(t["amount_due_kes"] for t in timeline), 2)
    months = len(timeline)
    cur_row = timeline[-1] if timeline else None
    cur_div = ((profile.division_override if profile and profile.division_override else "")
               or (cur_row["division"] if cur_row else ""))
    cur_name = ((profile.display_name if profile and profile.display_name else "")
                or (cur_row["raw_name"] if cur_row else ""))

    # Manual change / retag audit trail (matches pre-normalisation formats too)
    changes = [{
        "id": c.id, "entity_type": c.entity_type, "field": c.field,
        "prev_value": c.prev_value, "new_value": c.new_value, "actor": c.actor,
        "note": c.note, "created_at": str(c.created_at)[:19], "rolled_back": c.rolled_back,
    } for c in (db.query(ChangeLog)
                .filter(ChangeLog.org_id == org_id,
                        ChangeLog.entity_id.in_(_msisdn_variants(canonical)))
                .order_by(ChangeLog.created_at.desc()).limit(200).all())]

    return {
        "subscriber_number": canonical,
        "display_number": display_msisdn(canonical),
        "found": bool(timeline or profile),
        "as_of": as_of or (bills[-1].statement_iso if bills else ""),
        "available_periods": available_periods,
        "profile": None if not profile else {
            "display_name": profile.display_name,
            "division_override": profile.division_override, "tags": profile.tags,
            "device_type": profile.device_type, "tariff_override": profile.tariff_override,
            "notes": profile.notes, "is_active": profile.is_active,
            "first_seen_date": profile.first_seen_date, "last_seen_date": profile.last_seen_date,
        },
        "summary": {
            "status": status, "bills_present": months, "bills_total": len(bills),
            "first_seen": timeline[0]["statement_date"] if timeline
                          else (profile.first_seen_date if profile else ""),
            "last_seen": timeline[-1]["statement_date"] if timeline
                         else (profile.last_seen_date if profile else ""),
            "lifetime_spend": lifetime,
            "avg_monthly": round(lifetime / months, 2) if months else 0,
            "current_division": cur_div, "current_name": cur_name,
        },
        "timeline": timeline,
        "events": events,
        "changes": changes,
    }


def _bill_number_agg(bill_id: int, org_id: str, db: Session) -> dict:
    """Aggregate invoice lines for one bill into {number: {raw_name, division,
    amount_kes, cdr_count}} with the dominant line's name/division."""
    out: dict[str, dict] = {}
    for l in db.query(InvoiceLine).filter_by(org_id=org_id, bill_id=bill_id).all():
        a = out.get(l.subscriber_number)
        if a is None:
            a = out[l.subscriber_number] = {"raw_name": l.raw_name or "",
                "division": l.division or "", "amount_kes": 0.0,
                "cdr_count": 0, "_max": -1.0}
        a["amount_kes"] += l.amount_due_kes or 0
        a["cdr_count"] += l.cdr_count or 0
        if (l.amount_due_kes or 0) > a["_max"]:
            a["_max"] = l.amount_due_kes or 0
            a["raw_name"] = l.raw_name or a["raw_name"]
            a["division"] = l.division or a["division"]
    return out


def _compute_waste(org_id: str, db: Session) -> dict:
    """Shared waste/insight computation (reused by the endpoint and the report)."""
    bills = (db.query(BillUpload).filter_by(org_id=org_id)
             .order_by(BillUpload.statement_iso).all())
    if not bills:
        return {"bills_loaded": 0, "reference_period": None, "prev_period": None,
                "dormant_billed": [], "top_increases": [], "deactivated": [], "summary": {}}
    latest = bills[-1]
    prev = bills[-2] if len(bills) > 1 else None
    cur = _bill_number_agg(latest.id, org_id, db)
    pre = _bill_number_agg(prev.id, org_id, db) if prev else {}

    dormant_billed = sorted(
        [{"subscriber_number": n, "display_number": display_msisdn(n),
          "raw_name": a["raw_name"], "division": a["division"],
          "amount_kes": round(a["amount_kes"], 2)}
         for n, a in cur.items()
         if round(a["amount_kes"], 2) > 0 and a["cdr_count"] == 0],
        key=lambda x: x["amount_kes"], reverse=True)[:100]

    increases = []
    for n, a in cur.items():
        if n in pre:
            d = round(a["amount_kes"] - pre[n]["amount_kes"], 2)
            if d > 0:
                pct = round(d / pre[n]["amount_kes"] * 100, 1) if pre[n]["amount_kes"] else None
                increases.append({"subscriber_number": n, "display_number": display_msisdn(n),
                    "raw_name": a["raw_name"], "division": a["division"],
                    "prev_kes": round(pre[n]["amount_kes"], 2),
                    "curr_kes": round(a["amount_kes"], 2), "delta_kes": d, "delta_pct": pct})
    increases.sort(key=lambda x: x["delta_kes"], reverse=True)

    deactivated = sorted(
        [{"subscriber_number": n, "display_number": display_msisdn(n),
          "raw_name": a["raw_name"], "division": a["division"],
          "last_amount_kes": round(a["amount_kes"], 2)}
         for n, a in pre.items() if n not in cur],
        key=lambda x: x["last_amount_kes"], reverse=True)[:100]

    return {
        "bills_loaded": len(bills),
        "reference_period": latest.statement_date,
        "prev_period": prev.statement_date if prev else None,
        "dormant_billed": dormant_billed,
        "top_increases": increases[:25],
        "deactivated": deactivated,
        "summary": {
            "dormant_billed_count": len(dormant_billed),
            "dormant_billed_kes": round(sum(x["amount_kes"] for x in dormant_billed), 2),
            "deactivated_count": len(deactivated),
            "increase_count": len(increases),
            "total_increase_kes": round(sum(x["delta_kes"] for x in increases), 2),
        },
    }


@app.get("/api/orgs/{org_id}/waste")
def waste_insights(org_id: str, db: Session = Depends(get_db)):
    """Cost-saving signals as of the latest bill: lines billed but unused
    (dormant-but-billed), biggest month-over-month increases, and lines that
    dropped off (possible deactivations)."""
    return _compute_waste(org_id, db)


@app.get("/api/orgs/{org_id}/tags")
def list_tags(org_id: str, db: Session = Depends(get_db)):
    """Return all unique tags in use across subscribers for this org."""
    profiles = db.query(SubscriberProfile.tags).filter_by(org_id=org_id).all()
    tags: set[str] = set()
    for (t,) in profiles:
        if t:
            for tag in t.split(","):
                stripped = tag.strip()
                if stripped:
                    tags.add(stripped)
    return sorted(tags)


@app.get("/api/orgs/{org_id}/lifecycle")
def subscriber_lifecycle(org_id: str, db: Session = Depends(get_db)):
    """
    Compare subscriber lists across consecutive bills to detect:
    - New activations (first appearance)
    - Inactive lines (missing from latest bill vs prior)
    - Plan changes (tariff plan changed between bills)
    """
    bills = (
        db.query(BillUpload)
        .filter_by(org_id=org_id)
        .order_by(BillUpload.statement_iso)
        .all()
    )
    if len(bills) < 2:
        return {"bills_loaded": len(bills), "events": [],
                "message": "Load at least 2 bills to see lifecycle events."}

    events = []
    for i in range(1, len(bills)):
        prev_bill = bills[i - 1]
        curr_bill = bills[i]

        prev_lines = {
            l.subscriber_number: l
            for l in db.query(InvoiceLine).filter_by(bill_id=prev_bill.id).all()
        }
        curr_lines = {
            l.subscriber_number: l
            for l in db.query(InvoiceLine).filter_by(bill_id=curr_bill.id).all()
        }

        # New activations
        for sub, line in curr_lines.items():
            if sub not in prev_lines:
                events.append({
                    "event": "new_activation",
                    "subscriber_number": sub,
                    "name": line.raw_name,
                    "division": line.division,
                    "period": curr_bill.statement_date,
                    "detail": f"First appeared in {curr_bill.statement_date}",
                    "amount_kes": round(line.amount_due_kes, 2),
                })

        # Deactivations (in prev but not in curr)
        for sub, line in prev_lines.items():
            if sub not in curr_lines:
                events.append({
                    "event": "deactivation",
                    "subscriber_number": sub,
                    "name": line.raw_name,
                    "division": line.division,
                    "period": curr_bill.statement_date,
                    "detail": f"Last seen in {prev_bill.statement_date}",
                    "amount_kes": 0,
                })

        # Tariff plan changes
        for sub, curr_line in curr_lines.items():
            if sub in prev_lines:
                prev_tariff = prev_lines[sub].tariff_plan or ""
                curr_tariff = curr_line.tariff_plan or ""
                if prev_tariff and curr_tariff and prev_tariff != curr_tariff:
                    events.append({
                        "event": "plan_change",
                        "subscriber_number": sub,
                        "name": curr_line.raw_name,
                        "division": curr_line.division,
                        "period": curr_bill.statement_date,
                        "detail": f"{prev_tariff} → {curr_tariff}",
                        "amount_kes": round(curr_line.amount_due_kes, 2),
                    })

    events.sort(key=lambda e: (e["period"], e["event"]), reverse=True)
    return {
        "bills_loaded": len(bills),
        "events": events,
        "summary": {
            "new_activations": sum(1 for e in events if e["event"] == "new_activation"),
            "deactivations": sum(1 for e in events if e["event"] == "deactivation"),
            "plan_changes": sum(1 for e in events if e["event"] == "plan_change"),
        },
    }


# ═══════════════════════════════════════════════════════════════════════════════
# BUDGET & FORECASTING ENDPOINTS
# ═══════════════════════════════════════════════════════════════════════════════

@app.get("/api/orgs/{org_id}/budgets")
def get_budgets(org_id: str, db: Session = Depends(get_db)):
    rows = db.query(BudgetEntry).filter_by(org_id=org_id).order_by(
        BudgetEntry.period, BudgetEntry.division
    ).all()
    return [
        {
            "id": r.id, "division": r.division, "period": r.period,
            "budget_kes": r.budget_kes, "headcount": r.headcount,
        }
        for r in rows
    ]


@app.post("/api/orgs/{org_id}/budgets")
def upsert_budget(org_id: str, body: dict = Body(...), db: Session = Depends(get_db)):
    """Upsert a budget entry for a division+period. Body: {division, period, budget_kes, headcount}."""
    existing = db.query(BudgetEntry).filter_by(
        org_id=org_id, division=body["division"], period=body["period"]
    ).first()
    if existing:
        existing.budget_kes = float(body.get("budget_kes", 0))
        existing.headcount = int(body.get("headcount", 0))
    else:
        db.add(BudgetEntry(
            org_id=org_id,
            division=body["division"],
            period=body["period"],
            budget_kes=float(body.get("budget_kes", 0)),
            headcount=int(body.get("headcount", 0)),
        ))
    db.commit()
    return {"ok": True}


@app.delete("/api/orgs/{org_id}/budgets/{budget_id}")
def delete_budget(org_id: str, budget_id: int, db: Session = Depends(get_db)):
    db.query(BudgetEntry).filter_by(id=budget_id, org_id=org_id).delete()
    db.commit()
    return {"ok": True}


@app.post("/api/orgs/{org_id}/budgets/import-csv")
async def import_budget_csv(
    org_id: str,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    """
    Import budgets from CSV. Required columns: division, period (YYYY-MM), budget_kes.
    Optional: headcount.
    """
    content = (await file.read()).decode("utf-8", errors="replace")
    reader = csv.DictReader(io.StringIO(content))
    imported = 0
    errors = []
    for i, row in enumerate(reader, start=2):
        try:
            division = row.get("division", "").strip()
            period = row.get("period", "").strip()
            budget_kes = float(row.get("budget_kes", 0))
            headcount = int(row.get("headcount", 0) or 0)
            if not division or not period:
                errors.append(f"Row {i}: missing division or period")
                continue
            existing = db.query(BudgetEntry).filter_by(
                org_id=org_id, division=division, period=period
            ).first()
            if existing:
                existing.budget_kes = budget_kes
                existing.headcount = headcount
            else:
                db.add(BudgetEntry(
                    org_id=org_id, division=division, period=period,
                    budget_kes=budget_kes, headcount=headcount,
                ))
            imported += 1
        except (ValueError, KeyError) as e:
            errors.append(f"Row {i}: {e}")
    db.commit()
    return {"imported": imported, "errors": errors}


@app.get("/api/orgs/{org_id}/budget-vs-actual")
def budget_vs_actual(org_id: str, db: Session = Depends(get_db)):
    """
    For each bill period, compare actual division spend against budget.
    Returns per-period, per-division actuals + budget + variance + cost-per-head.
    """
    bills = (
        db.query(BillUpload)
        .filter_by(org_id=org_id)
        .order_by(BillUpload.statement_iso)
        .all()
    )
    budgets = db.query(BudgetEntry).filter_by(org_id=org_id).all()
    # Build budget lookup: {period: {division: {budget_kes, headcount}}}
    bmap: dict = {}
    for b in budgets:
        bmap.setdefault(b.period, {})[b.division] = {
            "budget_kes": b.budget_kes,
            "headcount": b.headcount,
        }

    result = []
    for bill in bills:
        # Normalise period to YYYY-MM
        sd = bill.statement_date or ""
        import re as _re
        m = _re.search(r"(\d{4})-(\d{2})", sd)
        if not m:
            m2 = _re.search(r"(\d{2})[/-](\d{4})", sd)
            period = f"{m2.group(2)}-{m2.group(1):0>2}" if m2 else sd[:7]
        else:
            period = f"{m.group(1)}-{m.group(2)}"

        lines = db.query(InvoiceLine).filter_by(bill_id=bill.id).all()
        div_actual: dict[str, float] = {}
        for line in lines:
            div_actual[line.division] = div_actual.get(line.division, 0) + line.amount_due_kes

        all_divs = set(div_actual.keys()) | set((bmap.get(period) or {}).keys())
        period_rows = []
        for div in sorted(all_divs):
            actual = round(div_actual.get(div, 0), 2)
            bentry = (bmap.get(period) or {}).get(div, {})
            budget = bentry.get("budget_kes", 0)
            headcount = bentry.get("headcount", 0)
            variance = round(actual - budget, 2) if budget else None
            pct_used = round((actual / budget) * 100, 1) if budget else None
            cost_per_head = round(actual / headcount, 2) if headcount else None
            period_rows.append({
                "division": div,
                "actual_kes": actual,
                "budget_kes": budget,
                "variance_kes": variance,
                "pct_of_budget": pct_used,
                "headcount": headcount,
                "cost_per_head": cost_per_head,
                "status": (
                    "over" if (pct_used or 0) > 100
                    else "warning" if (pct_used or 0) > 85
                    else "ok" if budget
                    else "no_budget"
                ),
            })

        result.append({
            "bill_id": bill.id,
            "period": period,
            "statement_date": bill.statement_date,
            "total_actual": round(sum(div_actual.values()), 2),
            "total_budget": round(sum(b.budget_kes for b in budgets if b.period == period), 2),
            "divisions": period_rows,
        })
    return result


@app.get("/api/orgs/{org_id}/forecast")
def spend_forecast(org_id: str, months_ahead: int = 3, db: Session = Depends(get_db)):
    """
    3-month linear regression forecast for total org spend and top divisions.
    Uses all loaded bill history.
    """
    bills = (
        db.query(BillUpload)
        .filter_by(org_id=org_id)
        .order_by(BillUpload.statement_iso)
        .all()
    )
    if len(bills) < 2:
        return {"forecast": [], "message": "Load at least 2 bills to generate a forecast."}

    # Total spend history
    totals = []
    for b in bills:
        amt = db.query(func.sum(InvoiceLine.amount_due_kes)).filter_by(bill_id=b.id).scalar() or 0
        totals.append(float(amt))

    n = len(totals)
    xs = list(range(n))
    x_mean = sum(xs) / n
    y_mean = sum(totals) / n
    denom = sum((x - x_mean) ** 2 for x in xs) or 1
    slope = sum((xs[i] - x_mean) * (totals[i] - y_mean) for i in range(n)) / denom
    intercept = y_mean - slope * x_mean

    # Forecast periods
    import re as _re
    last_date = bills[-1].statement_date or ""
    forecast_points = []
    for ahead in range(1, months_ahead + 1):
        predicted = max(0, intercept + slope * (n - 1 + ahead))
        # Estimate period label
        m = _re.search(r"(\d{4})-(\d{2})", last_date)
        if m:
            yr, mo = int(m.group(1)), int(m.group(2))
            mo += ahead
            while mo > 12:
                mo -= 12; yr += 1
            period_label = f"{yr}-{mo:02d}"
        else:
            period_label = f"T+{ahead}"
        forecast_points.append({
            "period": period_label,
            "predicted_kes": round(predicted, 2),
            "is_forecast": True,
        })

    # Confidence band (± 1 std-dev of residuals)
    residuals = [totals[i] - (intercept + slope * i) for i in range(n)]
    std_dev = (sum(r ** 2 for r in residuals) / max(n - 2, 1)) ** 0.5
    for pt in forecast_points:
        pt["lower_kes"] = round(max(0, pt["predicted_kes"] - std_dev), 2)
        pt["upper_kes"] = round(pt["predicted_kes"] + std_dev, 2)

    # Historical points for chart continuity
    history = []
    for i, b in enumerate(bills):
        history.append({
            "period": b.statement_date,
            "actual_kes": round(totals[i], 2),
            "trend_kes": round(intercept + slope * i, 2),
            "is_forecast": False,
        })

    return {
        "slope_kes_per_month": round(slope, 2),
        "history": history,
        "forecast": forecast_points,
        "std_dev": round(std_dev, 2),
    }


@app.get("/api/orgs/{org_id}/alerts")
def get_alerts(org_id: str, db: Session = Depends(get_db)):
    rows = db.query(SpendAlert).filter_by(org_id=org_id, is_active=True).all()
    return [
        {
            "id": r.id, "scope": r.scope, "scope_value": r.scope_value,
            "threshold_kes": r.threshold_kes, "is_active": r.is_active,
        }
        for r in rows
    ]


@app.post("/api/orgs/{org_id}/alerts")
def create_alert(org_id: str, body: dict = Body(...), db: Session = Depends(get_db)):
    alert = SpendAlert(
        org_id=org_id,
        scope=body.get("scope", "subscriber"),
        scope_value=body.get("scope_value", ""),
        threshold_kes=float(body.get("threshold_kes", 0)),
        is_active=True,
    )
    db.add(alert); db.commit()
    return {"id": alert.id}


@app.delete("/api/orgs/{org_id}/alerts/{alert_id}")
def delete_alert(org_id: str, alert_id: int, db: Session = Depends(get_db)):
    db.query(SpendAlert).filter_by(id=alert_id, org_id=org_id).delete()
    db.commit()
    return {"ok": True}


@app.get("/api/orgs/{org_id}/alert-breaches")
def alert_breaches(org_id: str, db: Session = Depends(get_db)):
    """
    For the latest bill, check which subscribers/divisions are breaching their alert thresholds.
    """
    latest_bill = (
        db.query(BillUpload)
        .filter_by(org_id=org_id)
        .order_by(BillUpload.statement_iso.desc())
        .first()
    )
    if not latest_bill:
        return []

    alerts = db.query(SpendAlert).filter_by(org_id=org_id, is_active=True).all()
    lines = db.query(InvoiceLine).filter_by(bill_id=latest_bill.id).all()

    # Build totals
    sub_totals: dict[str, float] = {}
    div_totals: dict[str, float] = {}
    grand_total = 0.0
    sub_names: dict[str, str] = {}
    for l in lines:
        sub_totals[l.subscriber_number] = sub_totals.get(l.subscriber_number, 0) + l.amount_due_kes
        div_totals[l.division] = div_totals.get(l.division, 0) + l.amount_due_kes
        grand_total += l.amount_due_kes
        sub_names[l.subscriber_number] = l.raw_name

    breaches = []
    for alert in alerts:
        if alert.scope == "subscriber":
            actual = sub_totals.get(alert.scope_value, 0)
            name = sub_names.get(alert.scope_value, alert.scope_value)
        elif alert.scope == "division":
            actual = div_totals.get(alert.scope_value, 0)
            name = alert.scope_value
        else:  # total
            actual = grand_total
            name = "Total Bill"

        if actual > alert.threshold_kes:
            breaches.append({
                "alert_id": alert.id,
                "scope": alert.scope,
                "scope_value": alert.scope_value,
                "name": name,
                "actual_kes": round(actual, 2),
                "threshold_kes": round(alert.threshold_kes, 2),
                "overage_kes": round(actual - alert.threshold_kes, 2),
                "pct_over": round(((actual - alert.threshold_kes) / alert.threshold_kes) * 100, 1),
            })
    return breaches




# ═══════════════════════════════════════════════════════════════════════════════
# DIVISION MANAGER ENDPOINTS
# ═══════════════════════════════════════════════════════════════════════════════

@app.get("/api/orgs/{org_id}/divisions")
def list_divisions(org_id: str, db: Session = Depends(get_db)):
    """Return the org's registered divisions, augmented with live spend from latest bill."""
    divs = db.query(Division).filter_by(org_id=org_id).order_by(Division.name).all()
    # Also surface divisions from InvoiceLines that may not yet be in registry
    line_divs = (
        db.query(InvoiceLine.division)
        .filter_by(org_id=org_id)
        .distinct()
        .all()
    )
    registered = {d.name for d in divs}
    for (ld,) in line_divs:
        if ld and ld not in registered:
            db.add(Division(org_id=org_id, name=ld))
            registered.add(ld)
    db.commit()
    result = db.query(Division).filter_by(org_id=org_id).order_by(Division.name).all()
    return [
        {"id": d.id, "name": d.name, "colour": d.colour, "description": d.description}
        for d in result
    ]


@app.post("/api/orgs/{org_id}/divisions")
def create_division(org_id: str, body: dict = Body(...), db: Session = Depends(get_db)):
    """Create a new division name for this org."""
    name = body.get("name", "").strip()
    if not name:
        raise HTTPException(400, "name is required")
    existing = db.query(Division).filter_by(org_id=org_id, name=name).first()
    if existing:
        return {"id": existing.id, "status": "exists"}
    div = Division(
        org_id=org_id, name=name,
        colour=body.get("colour", ""),
        description=body.get("description", ""),
    )
    db.add(div)
    db.commit()
    db.add(AuditLog(org_id=org_id, action="division.create", detail=f"name={name}"))
    db.commit()
    return {"id": div.id, "status": "created"}


@app.patch("/api/orgs/{org_id}/divisions/{div_id}")
def update_division(org_id: str, div_id: int, body: dict = Body(...),
                    db: Session = Depends(get_db)):
    """Rename a division and/or update colour / description."""
    div = db.query(Division).filter_by(id=div_id, org_id=org_id).first()
    if not div:
        raise HTTPException(404, "Division not found")
    old_name = div.name
    for k in ("name", "colour", "description"):
        if k in body:
            setattr(div, k, body[k])
    # If renamed, cascade to all InvoiceLines and SubscriberProfiles
    if "name" in body and body["name"] != old_name:
        new_name = body["name"].strip()
        if not new_name:
            raise HTTPException(400, "name cannot be empty")
        db.query(InvoiceLine).filter_by(org_id=org_id, division=old_name).update(
            {"division": new_name}
        )
        db.query(SubscriberProfile).filter_by(
            org_id=org_id, division_override=old_name
        ).update({"division_override": new_name})
        db.add(AuditLog(org_id=org_id, action="division.rename",
                        detail=f"{old_name!r} → {new_name!r}"))
    db.commit()
    return {"ok": True}


@app.delete("/api/orgs/{org_id}/divisions/{div_id}")
def delete_division(org_id: str, div_id: int, db: Session = Depends(get_db)):
    """Remove a division from the registry (does not affect existing line assignments)."""
    db.query(Division).filter_by(id=div_id, org_id=org_id).delete()
    db.commit()
    return {"ok": True}


# ── Bulk retag ────────────────────────────────────────────────────────────────

@app.post("/api/orgs/{org_id}/retag")
def bulk_retag(
    org_id: str,
    body: dict = Body(...),
    db: Session = Depends(get_db),
):
    """
    Bulk-reassign a division across all bills for this org.

    body fields:
      new_division : str   — target division name  (required)
      search       : str   — filter by subscriber name / number substring  (optional)
      from_division: str   — only retag lines currently in this division   (optional)
      subscriber_numbers: list[str] — explicit list of subscriber numbers  (optional)
      bill_id      : int   — limit to a single bill                        (optional)
      actor        : str   — who is making the change                      (optional)
      note         : str   — reason / notes                                (optional)
      dry_run      : bool  — preview without saving                        (optional)
    """
    new_div  = body.get("new_division", "").strip()
    if not new_div:
        raise HTTPException(400, "new_division is required")

    search      = body.get("search", "").strip()
    from_div    = body.get("from_division", "").strip()
    explicit    = body.get("subscriber_numbers", [])
    bill_id     = body.get("bill_id")
    actor       = body.get("actor", "user")
    note        = body.get("note", "")
    dry_run     = bool(body.get("dry_run", False))

    q = db.query(InvoiceLine).filter_by(org_id=org_id)
    if bill_id:
        q = q.filter_by(bill_id=int(bill_id))
    if from_div:
        q = q.filter_by(division=from_div)
    if explicit:
        q = q.filter(InvoiceLine.subscriber_number.in_(explicit))
    if search:
        safe = search.replace("%", r"\%").replace("_", r"\_")
        sig = _significant_digits(search)
        if sig:
            q = q.filter(InvoiceLine.subscriber_number.ilike(f"%{sig}%"))
        else:
            q = q.filter(
                InvoiceLine.raw_name.ilike(f"%{safe}%")
                | InvoiceLine.subscriber_number.ilike(f"%{safe}%")
            )

    lines = q.all()
    affected = [l for l in lines if l.division != new_div]

    if dry_run:
        return {
            "dry_run": True,
            "would_change": len(affected),
            "sample": [
                {"subscriber_number": l.subscriber_number,
                 "raw_name": l.raw_name,
                 "current_division": l.division,
                 "new_division": new_div}
                for l in affected[:20]
            ],
        }

    # Apply changes
    ensure_division(org_id, new_div, db)
    changed_subs: set[str] = set()
    for line in affected:
        _log_change(org_id, "invoice_line", line.subscriber_number,
                    "division", line.division, new_div, actor=actor, note=note, db=db)
        line.division = new_div
        changed_subs.add(line.subscriber_number)

    # Also update SubscriberProfile.division_override for persistence across future bills
    for sub_num in changed_subs:
        profile = db.query(SubscriberProfile).filter_by(
            org_id=org_id, subscriber_number=sub_num
        ).first()
        if not profile:
            profile = SubscriberProfile(org_id=org_id, subscriber_number=sub_num)
            db.add(profile)
        if profile.division_override != new_div:
            _log_change(org_id, "subscriber_profile", sub_num,
                        "division_override", profile.division_override, new_div,
                        actor=actor, note=note, db=db)
        profile.division_override = new_div

    db.add(AuditLog(
        org_id=org_id, action="division.bulk_retag",
        detail=f"new={new_div!r}, changed={len(affected)}, subs={len(changed_subs)}, "
               f"search={search!r}, from={from_div!r}",
    ))
    db.commit()
    logger.info("Bulk retag: org=%s new=%r changed=%d", org_id, new_div, len(affected))
    return {
        "changed_lines": len(affected),
        "changed_subscribers": len(changed_subs),
        "new_division": new_div,
    }


# ═══════════════════════════════════════════════════════════════════════════════
# CHANGE LOG + ROLLBACK ENDPOINTS
# ═══════════════════════════════════════════════════════════════════════════════

@app.get("/api/orgs/{org_id}/changes")
def get_change_log(
    org_id: str,
    entity_type: str = "",
    field: str = "",
    subscriber_number: str = "",
    limit: int = 100,
    include_rolled_back: bool = False,
    db: Session = Depends(get_db),
):
    """Return the detailed change log with rollback status."""
    q = db.query(ChangeLog).filter_by(org_id=org_id)
    if not include_rolled_back:
        q = q.filter_by(rolled_back=False)
    if entity_type:
        q = q.filter_by(entity_type=entity_type)
    if field:
        q = q.filter_by(field=field)
    if subscriber_number:
        q = q.filter_by(entity_id=subscriber_number)
    rows = q.order_by(ChangeLog.created_at.desc()).limit(min(limit, 1000)).all()
    return [
        {
            "id": r.id,
            "entity_type": r.entity_type,
            "entity_id": r.entity_id,
            "field": r.field,
            "prev_value": r.prev_value,
            "new_value": r.new_value,
            "actor": r.actor,
            "note": r.note,
            "created_at": str(r.created_at)[:19],
            "rolled_back": r.rolled_back,
            "rolled_back_at": str(r.rolled_back_at)[:19] if r.rolled_back_at else None,
        }
        for r in rows
    ]


@app.post("/api/orgs/{org_id}/changes/{change_id}/rollback")
def rollback_change(
    org_id: str,
    change_id: int,
    db: Session = Depends(get_db),
):
    """
    Revert a single change: restore prev_value to the entity.
    Works for both invoice_line division changes and subscriber_profile field changes.
    """
    entry = db.query(ChangeLog).filter_by(id=change_id, org_id=org_id).first()
    if not entry:
        raise HTTPException(404, "Change not found")
    if entry.rolled_back:
        raise HTTPException(409, "Already rolled back")

    if entry.entity_type == "invoice_line":
        # Update ALL invoice lines for this subscriber (across all bills in org)
        db.query(InvoiceLine).filter_by(
            org_id=org_id,
            subscriber_number=entry.entity_id,
            division=entry.new_value,
        ).update({"division": entry.prev_value})

    elif entry.entity_type == "subscriber_profile":
        profile = db.query(SubscriberProfile).filter_by(
            org_id=org_id, subscriber_number=entry.entity_id
        ).first()
        if profile and entry.field in {
            "display_name", "division_override", "tags",
            "device_type", "notes", "tariff_override",
        }:
            setattr(profile, entry.field, entry.prev_value)

    entry.rolled_back = True
    entry.rolled_back_at = datetime.utcnow()
    db.add(AuditLog(
        org_id=org_id, action="change.rollback",
        detail=f"change_id={change_id} entity={entry.entity_id} "
               f"field={entry.field} restored={entry.prev_value!r}",
    ))
    db.commit()
    return {"ok": True, "restored_value": entry.prev_value}


@app.post("/api/orgs/{org_id}/changes/rollback-bulk")
def rollback_bulk(
    org_id: str,
    body: dict = Body(...),
    db: Session = Depends(get_db),
):
    """Roll back multiple change IDs in one request."""
    ids = body.get("change_ids", [])
    if not ids:
        raise HTTPException(400, "change_ids list is required")
    rolled = 0
    errors = []
    for cid in ids:
        entry = db.query(ChangeLog).filter_by(id=cid, org_id=org_id).first()
        if not entry:
            errors.append(f"id={cid} not found")
            continue
        if entry.rolled_back:
            errors.append(f"id={cid} already rolled back")
            continue
        if entry.entity_type == "invoice_line":
            db.query(InvoiceLine).filter_by(
                org_id=org_id,
                subscriber_number=entry.entity_id,
                division=entry.new_value,
            ).update({"division": entry.prev_value})
        elif entry.entity_type == "subscriber_profile":
            profile = db.query(SubscriberProfile).filter_by(
                org_id=org_id, subscriber_number=entry.entity_id
            ).first()
            if profile and entry.field in {
                "display_name", "division_override", "tags",
                "device_type", "notes", "tariff_override",
            }:
                setattr(profile, entry.field, entry.prev_value)
        entry.rolled_back = True
        entry.rolled_back_at = datetime.utcnow()
        rolled += 1
    db.add(AuditLog(org_id=org_id, action="change.rollback_bulk",
                    detail=f"rolled={rolled} of {len(ids)}"))
    db.commit()
    return {"rolled_back": rolled, "errors": errors}




@app.get("/api/health/resources")
def resource_stats():
    """System resource snapshot — used by the dashboard health indicator."""
    try:
        import psutil, os as _os
        cpu = psutil.cpu_percent(interval=0.1)
        mem = psutil.virtual_memory()
        proc = psutil.Process(_os.getpid())
        return {
            "ok": True,
            "cpu_pct": round(cpu, 1),
            "ram_total_mb": round(mem.total / 1024 / 1024),
            "ram_used_mb":  round(mem.used  / 1024 / 1024),
            "ram_pct":      round(mem.percent, 1),
            "proc_rss_mb":  round(proc.memory_info().rss / 1024 / 1024, 1),
            "warning": mem.percent > 80,
            "critical": mem.percent > 90,
        }
    except ImportError:
        return {"ok": False, "error": "psutil not installed — run: pip install psutil"}


# ── Graceful shutdown ─────────────────────────────────────────────────────────

@app.post("/api/shutdown")
def graceful_shutdown(delay: int = 1):
    """Trigger a clean server shutdown (used by the Stop button in the UI)."""
    import threading, signal, os as _os
    def _stop():
        import time; time.sleep(delay)
        logger.info("Graceful shutdown requested via API")
        _os.kill(_os.getpid(), signal.SIGTERM)
    threading.Thread(target=_stop, daemon=True).start()
    return {"ok": True, "message": "Server shutting down…"}


# ── Division search for retag UI ──────────────────────────────────────────────

@app.get("/api/orgs/{org_id}/retag-preview")
def retag_preview(
    org_id: str,
    search: str = "",
    from_division: str = "",
    bill_id: int = None,
    db: Session = Depends(get_db),
):
    """Preview which lines would be affected by a retag without committing."""
    q = db.query(InvoiceLine).filter_by(org_id=org_id)
    if bill_id:
        q = q.filter_by(bill_id=bill_id)
    if from_division:
        q = q.filter_by(division=from_division)
    if search:
        safe = search.replace("%", r"\%").replace("_", r"\_")
        sig = _significant_digits(search)
        if sig:
            q = q.filter(InvoiceLine.subscriber_number.ilike(f"%{sig}%"))
        else:
            q = q.filter(
                InvoiceLine.raw_name.ilike(f"%{safe}%")
                | InvoiceLine.subscriber_number.ilike(f"%{safe}%")
                | InvoiceLine.tariff_plan.ilike(f"%{safe}%")
            )
    rows = q.order_by(InvoiceLine.amount_due_kes.desc()).limit(200).all()
    return {
        "count": len(rows),
        "lines": [
            {
                "subscriber_number": r.subscriber_number,
                "raw_name": r.raw_name,
                "tariff_plan": r.tariff_plan,
                "division": r.division,
                "amount_due_kes": round(r.amount_due_kes, 2),
                "bill_id": r.bill_id,
            }
            for r in rows
        ],
    }

# ── Serve dashboard ──────────────────────────────────────────────────────────
@app.get("/", response_class=HTMLResponse)
def serve_dashboard():
    index = STATIC_DIR / "index.html"
    # Never cache the SPA shell, so updates take effect on the next load without
    # a manual hard-refresh.
    no_cache = {"Cache-Control": "no-cache, no-store, must-revalidate",
                "Pragma": "no-cache", "Expires": "0"}
    if index.exists():
        return HTMLResponse(index.read_text(encoding="utf-8"), headers=no_cache)
    return HTMLResponse("<h2>TelecomLens API ✓</h2><p><a href='/docs'>API docs</a></p>")
